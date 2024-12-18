import requests
from bs4 import BeautifulSoup
import json
import os
import sys
import datetime
import time
from QtStockPriceMainWindow import Ui_MainWindow  # 導入轉換後的 UI 類
from QtStockTradingEditDialog import Ui_Dialog as Ui_StockTradingDialog
from QtStockDividendEditDialog import Ui_Dialog as Ui_StockDividendDialog
from QtStockCapitalReductionEditDialog import Ui_Dialog as Ui_StockCapitalReductionDialog
from QtDuplicateOptionDialog import Ui_Dialog as Ui_DuplicateOptionDialog
from PySide6.QtWidgets import QApplication, QMainWindow, QDialog, QButtonGroup, QMessageBox, QStyledItemDelegate, QFileDialog, QHeaderView
from PySide6.QtGui import QStandardItemModel, QStandardItem, QIcon, QBrush
from PySide6.QtCore import Qt, QModelIndex, QRect, QSignalBlocker
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font
from enum import Enum, IntEnum
from decimal import Decimal

#打包指令
# cd D:\_2.code\PythonStockPrice   
# pyinstaller --hidden-import "babel.numbers" --add-data "icon;./icon" --onefile --noconsole StockPriceMainWindow.py
# pyinstaller --hidden-import "babel.numbers" --add-data "icon;./icon" --onefile --console StockPriceMainWindow.py
# pyinstaller --hidden-import "babel.numbers" --add-data "icon;./icon" --console StockPriceMainWindow.py

# 要把.ui檔變成.py
# cd D:\_2.code\PythonStockPrice
# pyside6-uic QtStockPriceMainWindow.ui -o QtStockPriceMainWindow.py
# pyside6-uic QtStockTradingEditDialog.ui -o QtStockTradingEditDialog.py
# pyside6-uic QtStockDividendEditDialog.ui -o QtStockDividendEditDialog.py
# pyside6-uic QtStockCapitalReductionEditDialog.ui -o QtStockCapitalReductionEditDialog.py
# pyside6-uic QtDuplicateOptionDialog.ui -o QtDuplicateOptionDialog.py

g_user_dir = os.path.expanduser("~")  #開發模式跟打包模式下都是C:\Users\foxki
g_exe_dir = os.path.dirname(__file__) #開發模式下是D:\_2.code\PythonStockPrice #打包模式後是C:\Users\foxki\AppData\Local\Temp\_MEI60962 最後那個資料夾是暫時性的隨機名稱
g_exe2_dir = os.path.dirname( sys.executable ) #開發模式下是C:\Users\foxki\AppData\Local\Programs\Python\Python312 #打包模式後是:D:\_2.code\PythonStockPrice\dist
g_abs_dir = os.path.dirname( os.path.abspath(__file__) ) #開發模式下是D:\_2.code\PythonStockPrice #打包模式後是C:\Users\foxki\AppData\Local\Temp\_MEI60962 最後那個資料夾是暫時性的隨機名稱
print( "g_user_dir :" + g_user_dir ) #開發模式下是C:\Users\foxki
print( "g_exe_dir :" + g_exe_dir ) #開發模式下是D:\_2.code\PythonStockPrice #打包模式後是C:\Users\foxki\AppData\Local\Temp\_MEI60962 最後那個資料夾是暫時性的隨機名稱
print( "g_exe2_dir :" + g_exe2_dir ) #開發模式下是C:\Users\foxki\AppData\Local\Programs\Python\Python312 #打包模式後是:D:\_2.code\PythonStockPrice\dist
print( "g_abs_dir :" + g_abs_dir ) #開發模式下是D:\_2.code\PythonStockPrice #打包模式後是C:\Users\foxki\AppData\Local\Temp\_MEI60962 最後那個資料夾是暫時性的隨機名稱



g_list_stock_list_table_horizontal_header = [ '總成本', '庫存股數', '平均成本', '今日股價', '淨值', '損益', '匯出', '刪除' ]
if getattr( sys, 'frozen', False ):
    # PyInstaller 打包後執行時
    g_exe_root_dir = os.path.dirname(__file__) #C:\Users\foxki\AppData\Local\Temp\_MEI60962
    g_data_dir = os.path.join( g_user_dir, "AppData", "Local", "FoxInfo" ) #C:\Users\foxki\AppData\Local\FoxInfo
else:
    # 正常執行 Python 腳本時
    g_exe_root_dir = os.path.dirname( os.path.abspath(__file__) )
    g_data_dir = g_exe_root_dir

window_icon_file_path = os.path.join( g_exe_root_dir, 'icon\\FoxInfo.png' ) 
edit_icon_file_path = os.path.join( g_exe_root_dir, 'icon\\Edit.svg' ) 
edit_icon = QIcon( edit_icon_file_path ) 
delete_icon_file_path = os.path.join( g_exe_root_dir, 'icon\\Delete.svg' ) 
delete_icon = QIcon( delete_icon_file_path ) 
export_icon_file_path = os.path.join( g_exe_root_dir, 'icon\\Export.svg' ) 
export_icon = QIcon( export_icon_file_path ) 
g_trading_data_json_file_path = os.path.join( g_data_dir, 'StockInventory', 'TradingData.json' )
g_UISetting_file_path = os.path.join( g_data_dir, 'StockInventory', 'UISetting.config' )
g_stock_number_file_path = os.path.join( g_data_dir, 'StockInventory', 'StockNumber.txt' )
g_stock_price_file_path = os.path.join( g_data_dir, 'StockInventory', 'StockPrice.txt' )

class CenterIconDelegate( QStyledItemDelegate ):
    def paint( self, painter, option, index ):
        # 获取单元格数据
        icon = index.data( Qt.DecorationRole )  # 获取图标
        
        # 如果有图标
        if icon:
            rect = option.rect  # 单元格的绘制区域
            size = icon.actualSize( rect.size() ) * 0.7  # 图标实际尺寸
            
            # 计算居中位置
            x = rect.x() + ( rect.width() - size.width() ) // 2
            y = rect.y() + ( rect.height() - size.height() ) // 2
            target_rect = QRect( x, y, size.width(), size.height() )
            
            # 绘制图标
            icon.paint( painter, target_rect, Qt.AlignCenter )
        else:
            # 如果没有图标，使用默认绘制方法
            super().paint( painter, option, index )

class TradingType( IntEnum ):
    TEMPLATE = 0
    SELL = 1
    BUY = 2
    DIVIDEND = 3
    CAPITAL_REDUCTION = 4

class TradingData( Enum ):
    STOCK_NUMBER = 0
    TRADING_DATE = 1
    TRADING_TYPE = 2 # 0:賣出, 1:買進, 2:股利, 3:減資
    TRADING_PRICE = 3
    TRADING_COUNT = 4
    TRADING_FEE_DISCOUNT = 5
    STOCK_DIVIDEND_PER_SHARE = 6
    CASH_DIVIDEND_PER_SHARE = 7
    IS_REQUIRED_EXTRA_INSURANCE_FEE = 8
    CAPITAL_REDUCTION_PER_SHARE = 9
    SORTED_INDEX = 10 #不會記錄
    TRADING_VALUE = 11 #不會記錄
    TRADING_FEE = 12 #不會記錄
    TRADING_TAX = 13 #不會記錄
    TRADING_COST = 14 #不會記錄
    STOCK_DIVIDEND_GAIN = 15 #不會記錄
    CASH_DIVIDEND_GAIN = 16 #不會記錄
    EXTRA_INSURANCE_FEE = 17 #不會記錄
    ACCUMULATED_COST = 18 #不會記錄
    ACCUMULATED_INVENTORY = 19 #不會記錄
    AVERAGE_COST = 20 #不會記錄
    AUTO_DIVIDEND = 21 #不會記錄

class TradingCost( Enum ):
    TRADING_VALUE = 0
    TRADING_FEE = 1
    TRADING_TAX = 2
    TRADING_TOTAL_COST = 3

class Utility():
    def compute_cost( e_trading_type, f_trading_price, n_trading_count, f_trading_fee_discount, b_etf, b_daying_trading ):
        f_trading_price = Decimal( str( f_trading_price ) )#原本10.45 * 100000 = 1044999.999999999 然後取 int 就變成1044999，所以改用Decimal
        n_trading_count = Decimal( str( n_trading_count ) )
        f_trading_fee_discount = Decimal( str( f_trading_fee_discount ) )
        dict_result = {}
        if e_trading_type == TradingType.BUY or e_trading_type == TradingType.SELL:
            n_trading_value = int( f_trading_price * n_trading_count )
            n_trading_fee = int( n_trading_value * Decimal( '0.001425' ) * f_trading_fee_discount )
            if n_trading_fee < 20 and n_trading_fee != 0:
                n_trading_fee = 20
            if e_trading_type == TradingType.SELL:
                if b_etf:
                    n_trading_tax = int( n_trading_value * Decimal( '0.001' ) )
                elif b_daying_trading:
                    n_trading_tax = int( n_trading_value * Decimal( '0.0015' ) )
                else:
                    n_trading_tax = int( n_trading_value * Decimal( '0.003' ) )
            else:
                n_trading_tax = 0

            dict_result[ TradingCost.TRADING_VALUE ] = n_trading_value
            dict_result[ TradingCost.TRADING_FEE ] = n_trading_fee
            dict_result[ TradingCost.TRADING_TAX ] = n_trading_tax
            if e_trading_type == TradingType.BUY:
                dict_result[ TradingCost.TRADING_TOTAL_COST ] = n_trading_value + n_trading_fee + n_trading_tax
            else:
                dict_result[ TradingCost.TRADING_TOTAL_COST ] = n_trading_value - n_trading_fee - n_trading_tax
        else:   
            dict_result[ TradingCost.TRADING_VALUE ] = 0
            dict_result[ TradingCost.TRADING_FEE ] = 0
            dict_result[ TradingCost.TRADING_TAX ] = 0
            dict_result[ TradingCost.TRADING_TOTAL_COST ] = 0
        return dict_result

    def generate_trading_data( str_stock_number,            #股票代碼
                               str_trading_date,            #交易日期
                               e_trading_type,              #交易種類
                               f_trading_price,             #交易價格
                               n_trading_count,             #交易股數
                               f_trading_fee_discount,      #手續費折扣
                               f_stock_dividend_per_share,  #每股股票股利
                               f_cash_dividend_per_share,   #每股現金股利
                               b_extra_insurance_fee,       #是否需扣除補充保費
                               f_capital_reduction_per_share ): #每股減資金額
        dict_trading_data = {}
        dict_trading_data[ TradingData.STOCK_NUMBER ] = str_stock_number
        dict_trading_data[ TradingData.TRADING_DATE ] = str_trading_date
        dict_trading_data[ TradingData.TRADING_TYPE ] = e_trading_type
        dict_trading_data[ TradingData.TRADING_PRICE ] = f_trading_price
        dict_trading_data[ TradingData.TRADING_COUNT ] = n_trading_count
        dict_trading_data[ TradingData.TRADING_FEE_DISCOUNT ] = f_trading_fee_discount
        dict_trading_data[ TradingData.STOCK_DIVIDEND_PER_SHARE ] = f_stock_dividend_per_share
        dict_trading_data[ TradingData.CASH_DIVIDEND_PER_SHARE ] = f_cash_dividend_per_share
        dict_trading_data[ TradingData.IS_REQUIRED_EXTRA_INSURANCE_FEE ] = b_extra_insurance_fee
        dict_trading_data[ TradingData.CAPITAL_REDUCTION_PER_SHARE ] = f_capital_reduction_per_share
        return dict_trading_data

class ImportDataDuplicateOptionDialog( QDialog ):
    def __init__( self, parent = None ):
        super().__init__( parent )

        self.ui = Ui_DuplicateOptionDialog()
        self.ui.setupUi( self )
        window_icon = QIcon( window_icon_file_path ) 
        self.setWindowIcon( window_icon )
        self.ui.qtOkPushButton.clicked.connect( self.accept_data )
        self.ui.qtCancelPushButton.clicked.connect( self.cancel )
        self.b_overwrite = False

    def accept_data( self ):
        self.b_overwrite = self.ui.qtOverWriteRadioButton.isChecked()
        self.accept()
    
    def cancel( self ):
        self.reject()

class StockCapitalReductionEditDialog( QDialog ):
    def __init__( self, str_stock_number, str_stock_name, parent = None ):
        super().__init__( parent )

        self.ui = Ui_StockCapitalReductionDialog()
        self.ui.setupUi( self )
        window_icon = QIcon( window_icon_file_path ) 
        self.setWindowIcon( window_icon )

        self.ui.qtStockNumberLabel.setText( str_stock_number )
        self.ui.qtStockNameLabel.setText( str_stock_name )
        obj_current_date = datetime.datetime.today()
        self.ui.qtDateEdit.setDate( obj_current_date.date() )
        self.ui.qtDateEdit.setCalendarPopup( True )
        self.ui.qtOkButtonBox.accepted.connect( self.accept_data )
        self.ui.qtOkButtonBox.rejected.connect( self.cancel )
        self.dict_trading_data = {}

    def setup_trading_date( self, str_date ):
        self.ui.qtDateEdit.setDate( datetime.datetime.strptime( str_date, "%Y-%m-%d" ).date() )

    def setup_stock_capital_reduction( self, f_stock_capital_reduction_per_share ):
        self.ui.qtCapitalReductionDoubleSpinBox.setValue( f_stock_capital_reduction_per_share )

    def accept_data( self ):
        f_stock_capital_reduction_per_share = self.ui.qtCapitalReductionDoubleSpinBox.value()
        if f_stock_capital_reduction_per_share != 0:

            self.dict_trading_data = Utility.generate_trading_data( self.ui.qtStockNumberLabel.text(),                  #股票代碼
                                                                    self.ui.qtDateEdit.date().toString( "yyyy-MM-dd" ), #交易日期
                                                                    TradingType.CAPITAL_REDUCTION,                      #交易種類
                                                                    0,                                                  #交易價格                         
                                                                    0,                                                  #交易股數
                                                                    1,                                                  #手續費折扣                                   
                                                                    0,                                                  #每股股票股利
                                                                    0,                                                  #每股現金股利
                                                                    False,                                              #是否需扣除補充保費
                                                                    f_stock_capital_reduction_per_share )               #每股減資金額           
            self.accept()
        else:
            self.reject()
    
    def cancel( self ):
        self.reject()

class StockDividendEditDialog( QDialog ):
    def __init__( self, str_stock_number, str_stock_name, b_extra_insurance_fee, parent = None ):
        super().__init__( parent )

        self.ui = Ui_StockDividendDialog()
        self.ui.setupUi( self )
        window_icon = QIcon( window_icon_file_path ) 
        self.setWindowIcon( window_icon )

        self.ui.qtStockNumberLabel.setText( str_stock_number )
        self.ui.qtStockNameLabel.setText( str_stock_name )
        self.ui.qtExtraInsuranceFeeCheckBox.setChecked( b_extra_insurance_fee )
        obj_current_date = datetime.datetime.today()
        self.ui.qtDateEdit.setDate( obj_current_date.date() )
        self.ui.qtDateEdit.setCalendarPopup( True )
        self.ui.qtOkButtonBox.accepted.connect( self.accept_data )
        self.ui.qtOkButtonBox.rejected.connect( self.cancel )
        self.dict_trading_data = {}

    def setup_trading_date( self, str_date ):
        self.ui.qtDateEdit.setDate( datetime.datetime.strptime( str_date, "%Y-%m-%d" ).date() )

    def setup_stock_dividend( self, f_stock_dividend_per_share ):
        self.ui.qtStockDividendDoubleSpinBox.setValue( f_stock_dividend_per_share )

    def setup_cash_dividend( self, f_cash_dividend_per_share ):
        self.ui.qtCashDividendDoubleSpinBox.setValue( f_cash_dividend_per_share )

    def setup_extra_insurance_fee( self, b_extra_insurance_fee ):
        self.ui.qtExtraInsuranceFeeCheckBox.setChecked( b_extra_insurance_fee )

    def accept_data( self ):
        f_stock_dividend_per_share = self.ui.qtStockDividendDoubleSpinBox.value()
        f_cash_dividend_per_share = self.ui.qtCashDividendDoubleSpinBox.value()
        b_extra_insurance_fee = self.ui.qtExtraInsuranceFeeCheckBox.isChecked()
        if f_stock_dividend_per_share != 0 or f_cash_dividend_per_share != 0:

            self.dict_trading_data = Utility.generate_trading_data( self.ui.qtStockNumberLabel.text(),                  #股票代碼
                                                                    self.ui.qtDateEdit.date().toString( "yyyy-MM-dd" ), #交易日期
                                                                    TradingType.DIVIDEND,                               #交易種類
                                                                    0,                                                  #交易價格                         
                                                                    0,                                                  #交易股數
                                                                    1,                                                  #手續費折扣                                   
                                                                    f_stock_dividend_per_share,                         #每股股票股利
                                                                    f_cash_dividend_per_share,                          #每股現金股利
                                                                    b_extra_insurance_fee,                              #是否需扣除補充保費
                                                                    0 )                                                 #每股減資金額
            self.accept()
        else:
            self.reject()
    
    def cancel( self ):
        self.reject()

class StockTradingEditDialog( QDialog ):
    def __init__( self, str_stock_number, str_stock_name, b_etf, b_discount, f_discount_value, parent = None ):
        super().__init__( parent )

        self.ui = Ui_StockTradingDialog()
        self.ui.setupUi( self )
        window_icon = QIcon( window_icon_file_path ) 
        self.setWindowIcon( window_icon )

        self.ui.qtStockNumberLabel.setText( str_stock_number )
        self.ui.qtStockNameLabel.setText( str_stock_name )
        obj_current_date = datetime.datetime.today()
        self.ui.qtDateEdit.setDate( obj_current_date.date() )
        self.ui.qtDateEdit.setCalendarPopup( True )
        self.ui.qtDiscountCheckBox.setChecked( b_discount )
        self.ui.qtDiscountRateDoubleSpinBox.setValue( f_discount_value )

        self.ui.qtDiscountCheckBox.stateChanged.connect( self.on_discount_check_box_state_changed )
        self.ui.qtDiscountRateDoubleSpinBox.valueChanged.connect( self.compute_cost )
        self.ui.qtBuyRadioButton.toggled.connect( self.compute_cost )
        self.ui.qtSellRadioButton.toggled.connect( self.compute_cost )
        self.ui.qtCommonTradeRadioButton.toggled.connect( self.on_trading_type_changed )
        self.ui.qtOddTradeRadioButton.toggled.connect( self.on_trading_type_changed )
        self.ui.qtPriceDoubleSpinBox.valueChanged.connect( self.compute_cost )
        self.ui.qtCommonTradeCountSpinBox.valueChanged.connect( self.compute_cost )
        self.ui.qtOddTradeCountSpinBox.valueChanged.connect( self.compute_cost )
        self.ui.qtOkButtonBox.accepted.connect( self.accept_data )
        self.ui.qtOkButtonBox.rejected.connect( self.cancel )
        self.b_etf = b_etf

        self.dict_trading_data = {}

    def on_discount_check_box_state_changed( self, state ):
        if state == 2:
            self.ui.qtDiscountRateDoubleSpinBox.setEnabled( True )
        else:
            self.ui.qtDiscountRateDoubleSpinBox.setEnabled( False )

        self.compute_cost()

    def setup_trading_date( self, str_date ):
        self.ui.qtDateEdit.setDate( datetime.datetime.strptime( str_date, "%Y-%m-%d" ).date() )

    def setup_trading_type( self, e_trading_type ):
        if e_trading_type == TradingType.BUY:
            self.ui.qtBuyRadioButton.setChecked( True )
        else:
            self.ui.qtSellRadioButton.setChecked( True )

    def setup_trading_discount( self, f_discount_value ):
        if f_discount_value != 1:
            self.ui.qtDiscountCheckBox.setChecked( True )
            self.ui.qtDiscountRateDoubleSpinBox.setValue( f_discount_value * 10 )
        else:
            self.ui.qtDiscountCheckBox.setChecked( False )
            self.ui.qtDiscountRateDoubleSpinBox.setValue( 6 )

    def setup_trading_price( self, f_price ):
        self.ui.qtPriceDoubleSpinBox.setValue( f_price )

    def setup_trading_count( self, f_count ):
        if f_count % 1000 == 0:
            self.ui.qtCommonTradeRadioButton.setChecked( True )
            self.ui.qtCommonTradeCountSpinBox.setValue( f_count / 1000 )
            self.ui.qtOddTradeCountSpinBox.setValue( 0 )
        else:
            self.ui.qtOddTradeRadioButton.setChecked( True )
            self.ui.qtCommonTradeCountSpinBox.setValue( 0 )
            self.ui.qtOddTradeCountSpinBox.setValue( f_count )

    def accept_data( self ):

        if float( self.ui.qtTotalCostLineEdit.text().replace( ',', '' ) ) != 0:
            
            self.dict_trading_data = Utility.generate_trading_data( self.ui.qtStockNumberLabel.text(),                  #股票代碼
                                                                    self.ui.qtDateEdit.date().toString( "yyyy-MM-dd" ), #交易日期
                                                                    self.get_trading_type(),                            #交易種類
                                                                    self.ui.qtPriceDoubleSpinBox.value(),               #交易價格
                                                                    self.get_trading_count(),                           #交易股數
                                                                    self.get_trading_fee_discount(),                    #手續費折扣
                                                                    0,                                                  #每股股票股利
                                                                    0,                                                  #每股現金股利
                                                                    False,                                              #是否需扣除補充保費
                                                                    0 )                                                 #每股減資金額
            self.accept()
        else:
            self.reject()
    
    def cancel( self ):
        self.reject()

    def on_trading_type_changed( self ):
        if self.ui.qtCommonTradeRadioButton.isChecked():
            self.ui.qtCommonTradeCountSpinBox.setEnabled( True )
            self.ui.qtOddTradeCountSpinBox.setEnabled( False )
        else:
            self.ui.qtCommonTradeCountSpinBox.setEnabled( False )
            self.ui.qtOddTradeCountSpinBox.setEnabled( True )

        self.compute_cost()

    def get_trading_type( self ):
        if self.ui.qtBuyRadioButton.isChecked():
            return TradingType.BUY
        else:
            return TradingType.SELL

    def get_trading_count( self ):
        if self.ui.qtCommonTradeRadioButton.isChecked():
            return self.ui.qtCommonTradeCountSpinBox.value() * 1000
        else:
            return self.ui.qtOddTradeCountSpinBox.value()
        
    def get_trading_fee_discount( self ):
        if self.ui.qtDiscountCheckBox.isChecked():
            return self.ui.qtDiscountRateDoubleSpinBox.value() / 10
        else:
            return 1

    def compute_cost( self ):
        e_trading_type = self.get_trading_type()
        f_trading_price = self.ui.qtPriceDoubleSpinBox.value()
        n_trading_count = self.get_trading_count()
        f_trading_fee_discount = self.get_trading_fee_discount() 
        
        dict_result = Utility.compute_cost( e_trading_type, f_trading_price, n_trading_count, f_trading_fee_discount, self.b_etf, False )

        if e_trading_type == TradingType.BUY:
            self.ui.qtTradingValueLineEdit.setText( format( dict_result[ TradingCost.TRADING_VALUE ], ',' ) )
            self.ui.qtTotalCostLineEdit.setText( format( dict_result[ TradingCost.TRADING_TOTAL_COST ], ',' ) )
        elif e_trading_type == TradingType.SELL:
            self.ui.qtTradingValueLineEdit.setText( format( -dict_result[ TradingCost.TRADING_VALUE ], ',' ) )
            self.ui.qtTotalCostLineEdit.setText( format( -dict_result[ TradingCost.TRADING_TOTAL_COST ], ',' ) )
        self.ui.qtFeeLineEdit.setText( format( dict_result[ TradingCost.TRADING_FEE ], ',' ) )
        self.ui.qtTaxLineEdit.setText( format( dict_result[ TradingCost.TRADING_TAX ], ',' ) )

class MainWindow( QMainWindow ):
    def __init__(self):
        super( MainWindow, self ).__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi( self )  # 設置 UI
        window_icon = QIcon( window_icon_file_path ) 
        self.setWindowIcon( window_icon )

        delegate = CenterIconDelegate()
        self.stock_list_model = QStandardItemModel( 0, 0 )
        self.stock_list_model.setHorizontalHeaderLabels( g_list_stock_list_table_horizontal_header )
        self.ui.qtStockListTableView.verticalHeader().setSectionsMovable( True )
        self.ui.qtStockListTableView.verticalHeader().sectionMoved.connect( self.on_stock_list_table_vertical_header_section_moved )
        self.ui.qtStockListTableView.verticalHeader().sectionClicked.connect( self.on_stock_list_table_vertical_section_clicked )
        self.ui.qtStockListTableView.verticalHeader().setSectionResizeMode( QHeaderView.Fixed )
        self.ui.qtStockListTableView.horizontalHeader().sectionResized.connect( self.on_stock_list_table_horizontal_section_resized )
        self.ui.qtStockListTableView.setModel( self.stock_list_model )
        self.ui.qtStockListTableView.setItemDelegate( delegate )
        self.ui.qtStockListTableView.clicked.connect( lambda index: self.on_stock_list_table_item_clicked( index, self.stock_list_model ) )

        self.per_stock_trading_data_model = QStandardItemModel( 0, 0 ) 
        self.per_stock_trading_data_model.setVerticalHeaderLabels( self.get_trading_data_header() )
        self.ui.qtTradingDataTableView.setModel( self.per_stock_trading_data_model )
        self.ui.qtTradingDataTableView.setItemDelegate( delegate )
        self.ui.qtTradingDataTableView.horizontalHeader().hide()
        self.ui.qtTradingDataTableView.verticalHeader().setSectionResizeMode( QHeaderView.Fixed )
        self.ui.qtTradingDataTableView.clicked.connect( lambda index: self.on_trading_data_table_item_clicked( index, self.per_stock_trading_data_model ) )

        self.ui.qtStockInputLineEdit.textChanged.connect( self.on_stock_input_text_changed ) 

        self.ui.qtStockSelectComboBox.setVisible( False )
        self.ui.qtStockSelectComboBox.activated.connect( self.on_stock_select_combo_box_current_index_changed )
        self.ui.qtStockSelectComboBox.setStyleSheet( "QComboBox { combobox-popup: 0; }" )
        self.ui.qtStockSelectComboBox.setMaxVisibleItems( 10 )

        button_group_1 = QButtonGroup(self)
        button_group_1.addButton( self.ui.qtFromNewToOldRadioButton )
        button_group_1.addButton( self.ui.qtFromOldToNewRadioButton )
        self.ui.qtFromNewToOldRadioButton.setChecked( True )
        self.ui.qtFromNewToOldRadioButton.toggled.connect( self.on_change_display_mode )

        button_group_2 = QButtonGroup(self)
        button_group_2.addButton( self.ui.qtShowAllRadioButton )
        button_group_2.addButton( self.ui.qtShow10RadioButton )
        self.ui.qtShowAllRadioButton.setChecked( True )
        self.ui.qtShowAllRadioButton.toggled.connect( self.on_change_display_mode )

        button_group_3 = QButtonGroup(self)
        button_group_3.addButton( self.ui.qtShow1StockRadioButton )
        button_group_3.addButton( self.ui.qtShow1000StockRadioButton )
        self.ui.qtShow1StockRadioButton.setChecked( True )
        self.ui.qtShow1StockRadioButton.toggled.connect( self.on_change_display_mode )

        button_group_4 = QButtonGroup(self)
        button_group_4.addButton( self.ui.qtADYearRadioButton )
        button_group_4.addButton( self.ui.qtROCYearRadioButton )
        self.ui.qtADYearRadioButton.setChecked( True )
        self.ui.qtROCYearRadioButton.toggled.connect( self.on_change_display_mode )
        
        self.ui.qtAddStockPushButton.clicked.connect( self.on_add_stock_push_button_clicked )
        self.ui.qtDiscountCheckBox.stateChanged.connect( self.on_discount_check_box_state_changed )
        self.ui.qtDiscountRateDoubleSpinBox.valueChanged.connect( self.save_UI_state )
        self.ui.qtExtraInsuranceFeeCheckBox.stateChanged.connect( self.save_UI_state )

        self.ui.qtAddTradingDataPushButton.clicked.connect( self.on_add_trading_data_push_button_clicked )
        self.ui.qtAddDividendDataPushButton.clicked.connect( self.on_add_dividend_data_push_button_clicked )
        self.ui.qtAddCapitalReductionDataPushButton.clicked.connect( self.on_add_capital_reduction_data_push_button_clicked )
        self.ui.qtExportAllStockTradingDataPushButton.clicked.connect( self.on_export_all_to_excell_button_clicked )
        self.ui.qtExportSelectedStockTradingDataPushButton.clicked.connect( self.on_export_selected_to_excell_button_clicked )

        self.ui.qtActionExport.triggered.connect( self.on_export_trading_data_action_triggered )
        self.ui.qtActionImport.triggered.connect( self.on_import_trading_data_action_triggered )
        

        obj_current_date = datetime.datetime.today() - datetime.timedelta( days = 1 )
        str_date = obj_current_date.strftime('%Y%m%d')
        self.dict_all_company_number_to_name_and_type = self.download_all_company_stock_number( str_date )
        self.dict_all_company_number_to_price_info = self.download_day_stock_price( str_date )
        self.download_all_yearly_dividend_data( 2019, str_date )
        self.dict_auto_stock_yearly_dividned = self.load_all_yearly_dividend_data( 2019 )
        n_retry = 0
        while len( self.dict_all_company_number_to_price_info ) == 0:
            obj_current_date = obj_current_date - datetime.timedelta( days = 1 )
            n_weekday = obj_current_date.weekday()
            if n_weekday == 5 or n_weekday == 6:
                continue
            str_date = obj_current_date.strftime('%Y%m%d')
            self.dict_all_company_number_to_price_info = self.download_day_stock_price( str_date )
            n_retry += 1
            if n_retry > 30:
                break

        self.str_picked_stock_number = None
        self.dict_all_stock_trading_data = {}
        self.list_stock_list_column_width = []

        
        
        self.initialize()

    def on_stock_input_text_changed( self ):
        with QSignalBlocker( self.ui.qtStockSelectComboBox ), QSignalBlocker( self.ui.qtStockInputLineEdit ):
            self.ui.qtStockSelectComboBox.clear()
            str_stock_input = self.ui.qtStockInputLineEdit.text()
            if len( str_stock_input ) == 0:
                self.ui.qtStockSelectComboBox.setVisible( False )
                return
            self.ui.qtStockSelectComboBox.setVisible( True )

            for stock_number, list_stock_name_and_type in self.dict_all_company_number_to_name_and_type.items():
                str_stock_name = list_stock_name_and_type[ 0 ]
                if str_stock_input in stock_number or str_stock_input in str_stock_name:
                    self.ui.qtStockSelectComboBox.addItem( f"{stock_number} {str_stock_name}" )
            # self.ui.qtStockSelectComboBox.showPopup() #showPopup的話，focus會被搶走

            self.ui.qtStockInputLineEdit.setFocus()

    def on_stock_select_combo_box_current_index_changed( self, index ):
        str_stock_input = self.ui.qtStockSelectComboBox.currentText()
        self.ui.qtStockInputLineEdit.setText( str_stock_input )
        self.ui.qtStockSelectComboBox.setVisible( False )
        self.ui.qtStockInputLineEdit.setFocus()

    def on_add_stock_push_button_clicked( self ):
        str_stock_input = self.ui.qtStockInputLineEdit.text()
        self.ui.qtStockInputLineEdit.clear()
        str_first_four_chars = str_stock_input.split(" ")[0]
        if str_first_four_chars not in self.dict_all_company_number_to_name_and_type:
            b_find = False
            for stock_number, list_stock_name_and_type in self.dict_all_company_number_to_name_and_type.items():
                str_stock_name = list_stock_name_and_type[ 0 ]
                if str_first_four_chars == str_stock_name:
                    str_first_four_chars = stock_number
                    b_find = True
                    break
            if not b_find:
                QMessageBox.warning( self, "警告", "輸入的股票代碼不存在", QMessageBox.Ok )
                return
        
        if str_first_four_chars not in self.dict_all_stock_trading_data:
            dict_trading_data = Utility.generate_trading_data( str_first_four_chars, #股票代碼
                                                               "0001-01-01",         #交易日期
                                                               TradingType.TEMPLATE, #交易種類
                                                               0,                    #交易價格
                                                               0,                    #交易股數
                                                               1,                    #手續費折扣
                                                               0,                    #每股股票股利
                                                               0,                    #每股現金股利
                                                               False,                #是否需扣除補充保費
                                                               0 )                   #每股減資金額
            self.dict_all_stock_trading_data[ str_first_four_chars ] = [ dict_trading_data ]
            sorted_list = self.process_single_trading_data( str_first_four_chars )
            self.refresh_stock_list_table()
            self.auto_save_trading_data()

    def on_discount_check_box_state_changed( self, state ):
        if state == 2:
            self.ui.qtDiscountRateDoubleSpinBox.setEnabled( True )
        else:
            self.ui.qtDiscountRateDoubleSpinBox.setEnabled( False )

        self.save_UI_state()

    def on_change_display_mode( self ):
        if self.str_picked_stock_number != None:
            self.refresh_trading_data_table( self.dict_all_stock_trading_data[ self.str_picked_stock_number ] )

        self.save_UI_state()

    def on_add_trading_data_push_button_clicked( self ):
        if self.str_picked_stock_number is None:
            return
        str_stock_number = self.str_picked_stock_number
        list_stock_name_and_type = self.dict_all_company_number_to_name_and_type[ str_stock_number ]
        str_stock_name = list_stock_name_and_type[ 0 ]
        b_etf = self.dict_all_company_number_to_name_and_type[ str_stock_number ][ 1 ]
        dialog = StockTradingEditDialog( str_stock_number, str_stock_name, b_etf, self.ui.qtDiscountCheckBox.isChecked(), self.ui.qtDiscountRateDoubleSpinBox.value(), self )

        if dialog.exec():
            dict_trading_data = dialog.dict_trading_data
            self.dict_all_stock_trading_data[ str_stock_number ].append( dict_trading_data )
            sorted_list = self.process_single_trading_data( str_stock_number )
            self.refresh_stock_list_table()
            self.refresh_trading_data_table( sorted_list )
            self.auto_save_trading_data()

    def on_add_dividend_data_push_button_clicked( self ):
        if self.str_picked_stock_number is None:
            return
        str_stock_number = self.str_picked_stock_number
        list_stock_name_and_type = self.dict_all_company_number_to_name_and_type[ str_stock_number ]
        str_stock_name = list_stock_name_and_type[ 0 ]
        dialog = StockDividendEditDialog( str_stock_number, str_stock_name, self.ui.qtExtraInsuranceFeeCheckBox.isChecked(), self )

        if dialog.exec():
            dict_trading_data = dialog.dict_trading_data
            self.dict_all_stock_trading_data[ str_stock_number ].append( dict_trading_data )
            sorted_list = self.process_single_trading_data( str_stock_number )
            self.refresh_stock_list_table()
            self.refresh_trading_data_table( sorted_list )
            self.auto_save_trading_data()

    def on_add_capital_reduction_data_push_button_clicked( self ):
        if self.str_picked_stock_number is None:
            return
        str_stock_number = self.str_picked_stock_number
        list_stock_name_and_type = self.dict_all_company_number_to_name_and_type[ str_stock_number ]
        str_stock_name = list_stock_name_and_type[ 0 ]
        dialog = StockCapitalReductionEditDialog( str_stock_number, str_stock_name, self )

        if dialog.exec():
            dict_trading_data = dialog.dict_trading_data
            self.dict_all_stock_trading_data[ str_stock_number ].append( dict_trading_data )
            sorted_list = self.process_single_trading_data( str_stock_number )
            self.refresh_stock_list_table()
            self.refresh_trading_data_table( sorted_list )
            self.auto_save_trading_data()

    def on_stock_list_table_vertical_header_section_moved( self, n_logical_index, n_old_visual_index, n_new_visual_index ):
        list_stock_number = []
        for index_row,( key_stock_number, value ) in enumerate( self.dict_all_stock_trading_data.items() ):
            list_stock_number.append( key_stock_number )

        element = list_stock_number.pop( n_old_visual_index )
        list_stock_number.insert( n_new_visual_index, element )

        dict_all_stock_trading_data_new = {}
        for index_row, str_stock_number in enumerate( list_stock_number ):
            dict_all_stock_trading_data_new[ str_stock_number ] = self.dict_all_stock_trading_data[ str_stock_number ]


        self.dict_all_stock_trading_data = dict_all_stock_trading_data_new
        self.refresh_stock_list_table()
        self.auto_save_trading_data()

    def on_stock_list_table_vertical_section_clicked( self, n_logical_index ):
        header_text = self.stock_list_model.verticalHeaderItem( n_logical_index ).text()
        str_stock_number = header_text.split(" ")[0]

        if str_stock_number in self.dict_all_stock_trading_data:
            if str_stock_number != self.str_picked_stock_number:
                self.str_picked_stock_number = str_stock_number
                list_trading_data = self.dict_all_stock_trading_data[ str_stock_number ]
                self.refresh_trading_data_table( list_trading_data )
        self.update_button_enable_disable_status()

    def on_stock_list_table_horizontal_section_resized( self, n_logical_index, n_old_size, n_new_size ):
        self.list_stock_list_column_width[ n_logical_index ] = n_new_size
        self.save_UI_state()

    def on_stock_list_table_item_clicked( self, index: QModelIndex, table_model ):
        item = table_model.itemFromIndex( index )
        if item is not None:
            n_column = index.column()  # 獲取列索引
            n_row = index.row()  # 獲取行索引
            header_text = table_model.verticalHeaderItem( index.row() ).text()
            str_stock_number = header_text.split(" ")[0]
            
            if n_column == len( g_list_stock_list_table_horizontal_header ) - 1:#刪除
                result = self.show_message_box( "警告", f"確定要刪掉『{header_text}』的所有資料嗎?" )
                if result:
                    del self.dict_all_stock_trading_data[ str_stock_number ]
                    self.str_picked_stock_number = None
                    self.refresh_stock_list_table()
                    self.per_stock_trading_data_model.clear()
                    self.auto_save_trading_data()
            elif n_column == len( g_list_stock_list_table_horizontal_header ) - 2:#匯出
                file_path = self.open_save_json_file_dialog()
                if file_path:
                    dict_stock_trading_data = { str_stock_number: self.dict_all_stock_trading_data[ str_stock_number ] }
                    self.manual_save_trading_data( dict_stock_trading_data, file_path )
                
                if str_stock_number != self.str_picked_stock_number:
                    self.str_picked_stock_number = str_stock_number
                    list_trading_data = self.dict_all_stock_trading_data[ str_stock_number ]
                    self.refresh_trading_data_table( list_trading_data )
            elif str_stock_number in self.dict_all_stock_trading_data:
                if str_stock_number != self.str_picked_stock_number:
                    self.str_picked_stock_number = str_stock_number
                    list_trading_data = self.dict_all_stock_trading_data[ str_stock_number ]
                    self.refresh_trading_data_table( list_trading_data )

        self.update_button_enable_disable_status()

    def on_trading_data_table_item_clicked( self, index: QModelIndex, table_model ):
        item = table_model.itemFromIndex( index )
        if item is not None:
            n_column = index.column()  # 獲取列索引
            n_row = index.row()  # 獲取行索引
            list_trading_data = self.dict_all_stock_trading_data[ self.str_picked_stock_number ]

            if ( n_row == len( self.get_trading_data_header() ) - 2 or #編輯
                n_row == len( self.get_trading_data_header() ) - 1 ): #刪除

                if self.str_picked_stock_number is None:
                    return
                str_stock_number = self.str_picked_stock_number
                list_stock_name_and_type = self.dict_all_company_number_to_name_and_type[ str_stock_number ]
                str_stock_name = list_stock_name_and_type[ 0 ]

                hidden_data = table_model.data( index, Qt.UserRole )
                n_findindex = -1
                for index, dict_selected_data in enumerate( list_trading_data ):
                    if dict_selected_data[ TradingData.SORTED_INDEX ] == hidden_data:
                        n_findindex = index
                        break
                if n_findindex == -1:
                    return
                dict_selected_data = list_trading_data[ n_findindex ]

                if n_row == len( self.get_trading_data_header() ) - 2: #編輯
                    if dict_selected_data[ TradingData.TRADING_TYPE ] == TradingType.TEMPLATE:
                        return
                    if dict_selected_data[ TradingData.TRADING_TYPE ] == TradingType.BUY or dict_selected_data[ TradingData.TRADING_TYPE ] == TradingType.SELL:
                        b_etf = self.dict_all_company_number_to_name_and_type[ str_stock_number ][ 1 ]
                        dialog = StockTradingEditDialog( str_stock_number, str_stock_name, b_etf, self.ui.qtDiscountCheckBox.isChecked(), self.ui.qtDiscountRateDoubleSpinBox.value(), self )
                        dialog.setup_trading_date( dict_selected_data[ TradingData.TRADING_DATE ] )
                        dialog.setup_trading_type( dict_selected_data[ TradingData.TRADING_TYPE ] )
                        dialog.setup_trading_discount( dict_selected_data[ TradingData.TRADING_FEE_DISCOUNT ] )
                        dialog.setup_trading_price( dict_selected_data[ TradingData.TRADING_PRICE ] )
                        dialog.setup_trading_count( dict_selected_data[ TradingData.TRADING_COUNT ] )
                    elif dict_selected_data[ TradingData.TRADING_TYPE ] == TradingType.DIVIDEND:
                        dialog = StockDividendEditDialog( str_stock_number, str_stock_name, self.ui.qtExtraInsuranceFeeCheckBox.isChecked(), self )
                        dialog.setup_trading_date( dict_selected_data[ TradingData.TRADING_DATE ] )
                        dialog.setup_stock_dividend( dict_selected_data[ TradingData.STOCK_DIVIDEND_PER_SHARE ] )
                        dialog.setup_cash_dividend( dict_selected_data[ TradingData.CASH_DIVIDEND_PER_SHARE ] )
                        dialog.setup_extra_insurance_fee( dict_selected_data[ TradingData.IS_REQUIRED_EXTRA_INSURANCE_FEE ] )
                    elif dict_selected_data[ TradingData.TRADING_TYPE ] == TradingType.CAPITAL_REDUCTION:
                        dialog = StockCapitalReductionEditDialog( str_stock_number, str_stock_name, self )
                        dialog.setup_trading_date( dict_selected_data[ TradingData.TRADING_DATE ] )
                        dialog.setup_stock_capital_reduction( dict_selected_data[ TradingData.CAPITAL_REDUCTION_PER_SHARE ] )

                    if dialog.exec():
                        dict_trading_data = dialog.dict_trading_data
                        self.dict_all_stock_trading_data[ str_stock_number ][ n_findindex ] = dict_trading_data
                        sorted_list = self.process_single_trading_data( str_stock_number )
                        self.refresh_stock_list_table()
                        self.refresh_trading_data_table( sorted_list )
                        self.auto_save_trading_data()

                elif n_row == len( self.get_trading_data_header() ) - 1: #刪除
                    result = self.show_message_box( "警告", f"確定要刪掉這筆交易資料嗎?" )
                    if result:
                        del self.dict_all_stock_trading_data[ str_stock_number ][ n_findindex ]
                        sorted_list = self.process_single_trading_data( str_stock_number )
                        self.refresh_stock_list_table()
                        self.refresh_trading_data_table( sorted_list )
                        self.auto_save_trading_data()

    def export_trading_data_to_excel( self, worksheet, str_stock_number, str_stock_name ):
        str_title = str_stock_number + " " + str_stock_name
        worksheet.column_dimensions['A'].width = 30
        list_trading_data = self.dict_all_stock_trading_data[ str_stock_number ]

        data_index = 0
        n_row_start = 0
        for dict_per_trading_data in list_trading_data:
            
            e_trading_type = dict_per_trading_data[ TradingData.TRADING_TYPE ]
            if e_trading_type == TradingType.TEMPLATE:
                continue

            if data_index % 10 == 0:
                list_data_header = self.get_trading_data_header()
                list_data_header.insert( 0, str_title )

                n_row_start = int( ( len( list_data_header ) -2 + 1 ) * int( data_index / 10 ) )
                for index_row, str_header in enumerate( list_data_header ):
                    if index_row == len( list_data_header ) - 2:
                        break
                    worksheet.cell( row = n_row_start + index_row + 1, column = 1, value = str_header )
                    if index_row == 0:
                        worksheet.cell( row = n_row_start + index_row + 1, column = 1 ).font = Font( bold = True )
                index_column = 0

            worksheet.column_dimensions[ get_column_letter( index_column + 2 ) ].width = 15

            list_data = self.get_per_trading_data_text_list( dict_per_trading_data )
            list_data.insert( 0, "" )
            for index_row, str_data in enumerate( list_data ):
                str_data = str_data.replace( ',', '' )
                n_cell_row = n_row_start + index_row + 1
                n_cell_column = index_column + 2
                cell = worksheet.cell( row = n_cell_row, column = n_cell_column )
                if str_data == "買進":
                    color_fill = PatternFill( start_color = "DA9694", end_color = "DA9694", fill_type="solid")
                    cell.fill = color_fill
                elif str_data == "賣出":
                    color_fill = PatternFill( start_color = "76933C", end_color = "76933C", fill_type="solid")
                    cell.fill = color_fill
                elif str_data == "股利分配":
                    color_fill = PatternFill( start_color = "8DB4E2", end_color = "8DB4E2", fill_type="solid")
                    cell.fill = color_fill
                elif str_data == "減資":
                    color_fill = PatternFill( start_color = "B1A0C7", end_color = "B1A0C7", fill_type="solid")
                    cell.fill = color_fill

                str_cell = get_column_letter( n_cell_column ) + str( n_cell_row )
                if index_row == 1:
                    worksheet.cell( row = n_cell_row, column = n_cell_column, value = str_data )
                else:
                    try:
                        f_data= float( str_data )
                        if f_data.is_integer():
                            f_data = int( f_data )
                            worksheet[ str_cell ].number_format = "#,##0"  #顯示千位逗號
                        elif ( f_data * 10 ).is_integer():
                            worksheet[ str_cell ].number_format = "#,##0.0"
                        elif ( f_data * 100 ).is_integer():
                            worksheet[ str_cell ].number_format = "#,##0.00"
                        elif ( f_data * 1000 ).is_integer():
                            worksheet[ str_cell ].number_format = "#,##0.000"
                        worksheet.cell( row = n_cell_row, column = n_cell_column, value = f_data )
                    except ValueError:
                        worksheet.cell( row = n_cell_row, column = n_cell_column, value = str_data )

                worksheet[ str_cell ].alignment = Alignment( horizontal = "center", vertical = "center" )
            data_index += 1
            index_column += 1

    def on_export_selected_to_excell_button_clicked( self ):
        if self.str_picked_stock_number is None:
            return
        str_stock_number = self.str_picked_stock_number
        list_stock_name_and_type = self.dict_all_company_number_to_name_and_type[ str_stock_number ]
        str_stock_name = list_stock_name_and_type[ 0 ]
        file_path = self.open_save_excel_file_dialog()
        if file_path:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = str_stock_number + " " + str_stock_name
            self.export_trading_data_to_excel( worksheet, str_stock_number, str_stock_name )
            workbook.save( file_path )

    def on_export_all_to_excell_button_clicked( self ):
        file_path = self.open_save_excel_file_dialog()
        if file_path:
            workbook = Workbook()
            for index, ( key_stock_number, value ) in enumerate( self.dict_all_stock_trading_data.items() ):
                list_stock_name_and_type = self.dict_all_company_number_to_name_and_type[ key_stock_number ]
                str_stock_name = list_stock_name_and_type[ 0 ]
                str_tab_title = key_stock_number + " " + str_stock_name
                if index == 0:
                    worksheet = workbook.active
                    worksheet.title = str_tab_title
                else:
                    worksheet = workbook.create_sheet( str_tab_title, index )
                self.export_trading_data_to_excel( worksheet, key_stock_number, str_stock_name )
            workbook.save( file_path )

    def on_export_trading_data_action_triggered( self ):
        file_path = self.open_save_json_file_dialog()
        if file_path:
            self.manual_save_trading_data( self.dict_all_stock_trading_data, file_path )

    def on_import_trading_data_action_triggered( self ):
        file_path = self.open_load_json_file_dialog()
        if file_path:

            dict_all_stock_trading_data_new = {}
            self.load_trading_data( file_path, dict_all_stock_trading_data_new )
            b_duplicate = False
            for key_stock_number, value in dict_all_stock_trading_data_new.items():
                if key_stock_number in self.dict_all_stock_trading_data:
                    b_duplicate = True
                    break

            if b_duplicate:
                dialog = ImportDataDuplicateOptionDialog( self )
                if dialog.exec():
                    if dialog.b_overwrite:
                        self.dict_all_stock_trading_data.update( dict_all_stock_trading_data_new )
                        self.process_all_trading_data()
                        self.refresh_stock_list_table()
                        self.auto_save_trading_data()
                        if self.str_picked_stock_number != None:
                            self.refresh_trading_data_table( self.dict_all_stock_trading_data[ self.str_picked_stock_number ] )
                    else:
                        for key_stock_number, value in dict_all_stock_trading_data_new.items():
                            if key_stock_number in self.dict_all_stock_trading_data:
                                self.dict_all_stock_trading_data[ key_stock_number ] += value
                            else:
                                self.dict_all_stock_trading_data[ key_stock_number ] = value
                        self.process_all_trading_data()
                        self.refresh_stock_list_table()
                        self.auto_save_trading_data()
                        if self.str_picked_stock_number != None:
                            self.refresh_trading_data_table( self.dict_all_stock_trading_data[ self.str_picked_stock_number ] )

            else:
                self.dict_all_stock_trading_data.update( dict_all_stock_trading_data_new )
                self.process_all_trading_data()
                self.refresh_stock_list_table()
                self.auto_save_trading_data()
                if self.str_picked_stock_number != None:
                    self.refresh_trading_data_table( self.dict_all_stock_trading_data[ self.str_picked_stock_number ] )

    def open_load_json_file_dialog( self ):
        # 彈出讀取檔案對話框
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "匯入交易資料",     # 對話框標題
            "",                # 預設路徑
            "JSON (*.json);;"  # 檔案類型過濾
        )
        return file_path

    def open_save_json_file_dialog( self ):
        # 彈出儲存檔案對話框
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "匯出交易資料",     # 對話框標題
            "",                # 預設路徑
            "JSON (*.json);;"  # 檔案類型過濾
        )
        return file_path
    
    def open_save_excel_file_dialog( self ):
        # 彈出儲存檔案對話框
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save File",             # 對話框標題
            "",                      # 預設路徑
            "Excel 活頁簿 (*.xlsx);;All Files (*)"  # 檔案類型過濾
        )
        return file_path

    def show_message_box( self, str_title, str_message ):
        message_box = QMessageBox( self )
        message_box.setIcon( QMessageBox.Warning )  # 設置為警告圖示
        message_box.setWindowTitle( str_title )
        message_box.setText( str_message )

        # 添加自訂按鈕
        button_ok = message_box.addButton("確定", QMessageBox.AcceptRole)
        button_cancel = message_box.addButton("取消", QMessageBox.RejectRole)

        message_box.exec()

        if message_box.clickedButton() == button_ok:
            return True
        elif message_box.clickedButton() == button_cancel:
            return False

    def update_button_enable_disable_status( self ):
        if self.str_picked_stock_number is None:
            self.ui.qtAddTradingDataPushButton.setEnabled( False )
            self.ui.qtAddDividendDataPushButton.setEnabled( False )
            self.ui.qtAddCapitalReductionDataPushButton.setEnabled( False )
            self.ui.qtExportSelectedStockTradingDataPushButton.setEnabled( False )
        else:
            self.ui.qtAddTradingDataPushButton.setEnabled( True )
            self.ui.qtAddDividendDataPushButton.setEnabled( True )
            self.ui.qtAddCapitalReductionDataPushButton.setEnabled( True )
            self.ui.qtExportSelectedStockTradingDataPushButton.setEnabled( True )

    def save_UI_state( self ):
        # 確保目錄存在，若不存在則遞歸創建
        os.makedirs( os.path.dirname( g_UISetting_file_path ), exist_ok = True )

        with open( g_UISetting_file_path, 'w', encoding='utf-8' ) as f:
            f.write( "手續費折扣," + str( self.ui.qtDiscountCheckBox.isChecked() ) + '\n' )
            f.write( "手續費折數," + str( self.ui.qtDiscountRateDoubleSpinBox.value() ) + '\n' )
            f.write( "補充保費," + str( self.ui.qtExtraInsuranceFeeCheckBox.isChecked() ) + '\n' )
            f.write( "顯示排序," + str( self.ui.qtFromNewToOldRadioButton.isChecked() ) + '\n' )
            f.write( "顯示數量," + str( self.ui.qtShowAllRadioButton.isChecked() ) + '\n' )
            f.write( "顯示單位," + str( self.ui.qtShow1StockRadioButton.isChecked() ) + '\n' )
            f.write( "欄寬" )
            for i in range( len( self.list_stock_list_column_width ) ):
                f.write( f",{ self.list_stock_list_column_width[ i ] }" )
            f.write( "\n" )

    def load_UI_state( self ):
        with ( QSignalBlocker( self.ui.qtStockListTableView.horizontalHeader() ),
               QSignalBlocker( self.ui.qtDiscountCheckBox ),
               QSignalBlocker( self.ui.qtDiscountRateDoubleSpinBox ),
               QSignalBlocker( self.ui.qtExtraInsuranceFeeCheckBox ), 
               QSignalBlocker( self.ui.qtFromNewToOldRadioButton ),
               QSignalBlocker( self.ui.qtFromOldToNewRadioButton ), 
               QSignalBlocker( self.ui.qtShowAllRadioButton ), 
               QSignalBlocker( self.ui.qtShow10RadioButton ),
               QSignalBlocker( self.ui.qtShow1StockRadioButton ), 
               QSignalBlocker( self.ui.qtShow1000StockRadioButton ) ):

            if os.path.exists( g_UISetting_file_path ):
                with open( g_UISetting_file_path, 'r', encoding='utf-8' ) as f:
                    data = f.readlines()
                    for i, row in enumerate( data ):
                        row = row.strip().split( ',' )
                        if row[0] == "手續費折扣":
                            self.ui.qtDiscountCheckBox.setChecked( row[ 1 ] == 'True' )
                            self.ui.qtDiscountRateDoubleSpinBox.setEnabled( row[ 1 ] == 'True' )
                        elif row[0] == "手續費折數":
                            self.ui.qtDiscountRateDoubleSpinBox.setValue( float(row[ 1 ]) )
                        elif row[0] == "補充保費":
                            self.ui.qtExtraInsuranceFeeCheckBox.setChecked( row[ 1 ] == 'True' )
                        elif row[0] == "顯示排序":
                            if row[ 1 ] == 'True':
                                self.ui.qtFromNewToOldRadioButton.setChecked( True )
                            else:
                                self.ui.qtFromOldToNewRadioButton.setChecked( True )
                        elif row[0] == "顯示數量":
                            if row[ 1 ] == 'True':
                                self.ui.qtShowAllRadioButton.setChecked( True )
                            else:
                                self.ui.qtShow10RadioButton.setChecked( True )
                        elif row[0] == "顯示單位":
                            if row[ 1 ] == 'True':
                                self.ui.qtShow1StockRadioButton.setChecked( True )
                            else:
                                self.ui.qtShow1000StockRadioButton.setChecked( True )
                        elif row[0] == '欄寬':
                            self.list_stock_list_column_width = []
                            for i in range( 1, len( row ) ):
                                self.list_stock_list_column_width.append( int( row[ i ] ) )

    def process_single_trading_data( self, str_stock_number ):
        list_trading_data = self.dict_all_stock_trading_data[ str_stock_number ]
        b_etf = self.dict_all_company_number_to_name_and_type[ str_stock_number ][ 1 ]
        sorted_list = sorted( list_trading_data, key=lambda x: ( datetime.datetime.strptime( x[ TradingData.TRADING_DATE ], "%Y-%m-%d"), -x[ TradingData.TRADING_TYPE ] ) )

        str_current_date = datetime.datetime.today().strftime("%Y-%m-%d")

        if str_stock_number in self.dict_auto_stock_yearly_dividned:
            auto_list_dividend = self.dict_auto_stock_yearly_dividned[ str_stock_number ]
            auto_list_dividend = sorted( auto_list_dividend, key=lambda x: ( datetime.datetime.strptime( x[ TradingData.TRADING_DATE ], "%Y-%m-%d") ) )
            if len( sorted_list ) > 1:
                first_data = sorted_list[ 1 ]
                for index, auto_dividend_data in enumerate( auto_list_dividend ):
                    if auto_dividend_data[ TradingData.TRADING_DATE ] > first_data[ TradingData.TRADING_DATE ]:
                        if auto_dividend_data[ TradingData.TRADING_DATE ] > str_current_date:
                            break
                        # sorted_list.append( auto_dividend_data )
        sorted_list = sorted( sorted_list, key=lambda x: ( datetime.datetime.strptime( x[ TradingData.TRADING_DATE ], "%Y-%m-%d"), -x[ TradingData.TRADING_TYPE ] ) )

        n_accumulated_inventory = 0
        n_accumulated_cost = 0
        str_last_buying_date = ''
        n_last_buying_count = 0
        list_calibration_data = [] #因為若是已經沒有庫存股票，那麼股利分配或是減資的資料就不會被計算
        for index, item in enumerate( sorted_list ):
            item[ TradingData.SORTED_INDEX ] = index
            e_trading_type = item[ TradingData.TRADING_TYPE ]
            
            if e_trading_type == TradingType.BUY:
                f_trading_price = item[ TradingData.TRADING_PRICE ]
                n_trading_count = item[ TradingData.TRADING_COUNT ]
                f_trading_fee_discount = item[ TradingData.TRADING_FEE_DISCOUNT ]
                dict_result = Utility.compute_cost( e_trading_type, f_trading_price, n_trading_count, f_trading_fee_discount, b_etf, False )
                item[ TradingData.TRADING_VALUE ] = dict_result[ TradingCost.TRADING_VALUE ]
                item[ TradingData.TRADING_FEE ] = dict_result[ TradingCost.TRADING_FEE ]
                item[ TradingData.TRADING_TAX ] = dict_result[ TradingCost.TRADING_TAX ]
                item[ TradingData.EXTRA_INSURANCE_FEE ] = 0 
                n_per_trading_total_cost = item[ TradingData.TRADING_COST ] = dict_result[ TradingCost.TRADING_TOTAL_COST ]
                n_accumulated_inventory += n_trading_count
                n_accumulated_cost += n_per_trading_total_cost

                item[ TradingData.STOCK_DIVIDEND_GAIN ] = 0
                item[ TradingData.CASH_DIVIDEND_GAIN ] = 0
                str_buying_date = item[ TradingData.TRADING_DATE ]
                if str_last_buying_date == str_buying_date:
                    n_last_buying_count += n_trading_count
                else:
                    str_last_buying_date = str_buying_date
                    n_last_buying_count = n_trading_count
                list_calibration_data.append( item )
            elif e_trading_type == TradingType.SELL:
                str_selling_date = item[ TradingData.TRADING_DATE ]
                f_trading_price = item[ TradingData.TRADING_PRICE ]
                f_trading_fee_discount = item[ TradingData.TRADING_FEE_DISCOUNT ]
                n_trading_count = item[ TradingData.TRADING_COUNT ]

                if str_selling_date == str_last_buying_date: #賣出與買入同一天屬於當沖
                    if n_trading_count <= n_last_buying_count: #賣出數量小於或等於買入數量，表示全部賣出數量都可視為當沖
                        dict_result = Utility.compute_cost( e_trading_type, f_trading_price, n_trading_count, f_trading_fee_discount, b_etf, True )
                        item[ TradingData.TRADING_VALUE ] = dict_result[ TradingCost.TRADING_VALUE ]
                        item[ TradingData.TRADING_FEE ] = dict_result[ TradingCost.TRADING_FEE ]
                        item[ TradingData.TRADING_TAX ] = dict_result[ TradingCost.TRADING_TAX ]
                        n_per_trading_total_cost = item[ TradingData.TRADING_COST ] = dict_result[ TradingCost.TRADING_TOTAL_COST ]
                        n_accumulated_cost -= n_per_trading_total_cost
                        n_accumulated_inventory -= n_trading_count
                        n_last_buying_count -= n_trading_count
                    else: #賣出數量大於買入數量，表示只有部分數量都可視為當沖
                        n_trading_count_1 = n_last_buying_count
                        dict_result = Utility.compute_cost( e_trading_type, f_trading_price, n_trading_count_1, f_trading_fee_discount, b_etf, True )#這部分是當沖
                        n_trading_value_1 = dict_result[ TradingCost.TRADING_VALUE ]
                        n_trading_fee_1 = dict_result[ TradingCost.TRADING_FEE ]
                        n_trading_tax_1 = dict_result[ TradingCost.TRADING_TAX ]
                        n_trading_total_cost_1 = dict_result[ TradingCost.TRADING_TOTAL_COST ]

                        n_trading_count_2 = n_trading_count - n_last_buying_count
                        dict_result = Utility.compute_cost( e_trading_type, f_trading_price, n_trading_count_2, f_trading_fee_discount, b_etf, False )#這部分不是當沖
                        n_trading_value_2 = dict_result[ TradingCost.TRADING_VALUE ]
                        n_trading_fee_2 = dict_result[ TradingCost.TRADING_FEE ]
                        n_trading_tax_2 = dict_result[ TradingCost.TRADING_TAX ]
                        n_trading_total_cost_2 = dict_result[ TradingCost.TRADING_TOTAL_COST ]

                        item[ TradingData.TRADING_VALUE ] = n_trading_value_1 + n_trading_value_2
                        item[ TradingData.TRADING_FEE ] = n_trading_fee_1 + n_trading_fee_2
                        item[ TradingData.TRADING_TAX ] = n_trading_tax_1 + n_trading_tax_2
                        n_per_trading_total_cost = item[ TradingData.TRADING_COST ] = n_trading_total_cost_1 + n_trading_total_cost_2
                        n_accumulated_cost -= n_per_trading_total_cost
                        n_accumulated_inventory -= n_trading_count

                        n_last_buying_count = 0
                else:
                    dict_result = Utility.compute_cost( e_trading_type, f_trading_price, n_trading_count, f_trading_fee_discount, b_etf, False )
                    item[ TradingData.TRADING_VALUE ] = dict_result[ TradingCost.TRADING_VALUE ]
                    item[ TradingData.TRADING_FEE ] = dict_result[ TradingCost.TRADING_FEE ]
                    item[ TradingData.TRADING_TAX ] = dict_result[ TradingCost.TRADING_TAX ]
                    n_per_trading_total_cost = item[ TradingData.TRADING_COST ] = dict_result[ TradingCost.TRADING_TOTAL_COST ]
                    n_accumulated_cost -= n_per_trading_total_cost
                    n_accumulated_inventory -= n_trading_count

                item[ TradingData.EXTRA_INSURANCE_FEE ] = 0
                item[ TradingData.STOCK_DIVIDEND_GAIN ] = 0
                item[ TradingData.CASH_DIVIDEND_GAIN ] = 0
                list_calibration_data.append( item )
            elif e_trading_type == TradingType.DIVIDEND:
                if n_accumulated_inventory > 0: #沒有庫存就不用算股利了
                    item[ TradingData.TRADING_VALUE ] = 0
                    item[ TradingData.TRADING_TAX ] = 0
                    item[ TradingData.TRADING_COST ] = 0

                    n_stock_dividend_gain = int( Decimal( str( item[ TradingData.STOCK_DIVIDEND_PER_SHARE ] ) ) * Decimal( str( n_accumulated_inventory ) ) / Decimal( '10' ) ) #f_stock_dividend_gain單位為股 除以10是因為票面額10元
                    item[ TradingData.STOCK_DIVIDEND_GAIN ] = n_stock_dividend_gain
                    n_cash_dividend_gain = int( Decimal( str(item[ TradingData.CASH_DIVIDEND_PER_SHARE ] ) ) * Decimal( str( n_accumulated_inventory ) ) )
                    n_accumulated_inventory += n_stock_dividend_gain
                    
                    if n_cash_dividend_gain > 10:
                        item[ TradingData.CASH_DIVIDEND_GAIN ] = n_cash_dividend_gain
                        item[ TradingData.TRADING_FEE ] = 10
                        if item[ TradingData.IS_REQUIRED_EXTRA_INSURANCE_FEE ] and n_cash_dividend_gain >= 20000:
                            n_extra_insurance_fee = int( Decimal( str( n_cash_dividend_gain ) ) * Decimal( str( '0.0211' ) ) )
                        else:
                            n_extra_insurance_fee = 0
                        item[ TradingData.EXTRA_INSURANCE_FEE ] = n_extra_insurance_fee
                        n_accumulated_cost = n_accumulated_cost - n_cash_dividend_gain + 10 + n_extra_insurance_fee
                    else:
                        item[ TradingData.CASH_DIVIDEND_GAIN ] = 0
                        item[ TradingData.TRADING_FEE ] = 0
                        item[ TradingData.EXTRA_INSURANCE_FEE ] = 0 
                    list_calibration_data.append( item )
            elif e_trading_type == TradingType.CAPITAL_REDUCTION:
                if n_accumulated_inventory > 0: #沒有庫存就不用算減資了
                    item[ TradingData.TRADING_PRICE ] = -item[ TradingData.CAPITAL_REDUCTION_PER_SHARE ]
                    item[ TradingData.TRADING_COUNT ] = n_accumulated_inventory
                    item[ TradingData.TRADING_VALUE ] = -int( n_accumulated_inventory * item[ TradingData.CAPITAL_REDUCTION_PER_SHARE ] )
                    item[ TradingData.TRADING_FEE ] = 0
                    item[ TradingData.TRADING_TAX ] = 0
                    item[ TradingData.EXTRA_INSURANCE_FEE ] = 0 
                    item[ TradingData.TRADING_COST ] = 0
                    item[ TradingData.STOCK_DIVIDEND_GAIN ] = 0
                    item[ TradingData.CASH_DIVIDEND_GAIN ] = 0
                    n_accumulated_cost = n_accumulated_cost - int( Decimal( str( n_accumulated_inventory ) ) * Decimal( str( item[ TradingData.CAPITAL_REDUCTION_PER_SHARE ] ) ) )
                    n_accumulated_inventory = int( Decimal( str( n_accumulated_inventory ) ) * ( Decimal( str( '10' ) ) - Decimal( str( item[ TradingData.CAPITAL_REDUCTION_PER_SHARE ] ) ) ) / Decimal( str( '10' ) ) )
                    list_calibration_data.append( item )
            item[ TradingData.ACCUMULATED_COST ] = n_accumulated_cost
            item[ TradingData.ACCUMULATED_INVENTORY ] = n_accumulated_inventory
            item[ TradingData.AVERAGE_COST ] = n_accumulated_cost / n_accumulated_inventory if n_accumulated_inventory != 0 else 0

        self.dict_all_stock_trading_data[ str_stock_number ] = list_calibration_data
        return list_calibration_data

    def process_all_trading_data( self ):
        for key_stock_number, value_list_trading_data in self.dict_all_stock_trading_data.items():
            self.process_single_trading_data( key_stock_number )

    def auto_save_trading_data( self ):
        self.manual_save_trading_data( self.dict_all_stock_trading_data, g_trading_data_json_file_path )

    def manual_save_trading_data( self, dict_stock_trading_data, file_path ):
        list_all_company_trading_data = []
        dict_all_trading_data = {}
        export_data = []
        for key, value in dict_stock_trading_data.items():
            for item in value:
                if item[ TradingData.AUTO_DIVIDEND ] == True:
                    continue
                dict_per_trading_data = {}
                dict_per_trading_data[ "stock_number" ] = item[ TradingData.STOCK_NUMBER ]
                dict_per_trading_data[ "trading_date" ] = item[ TradingData.TRADING_DATE ]
                dict_per_trading_data[ "trading_type" ] = int( item[ TradingData.TRADING_TYPE ].value )
                if item[ TradingData.TRADING_TYPE ] == TradingType.CAPITAL_REDUCTION: #CAPITAL_REDUCTION 為了顯示，所以需要寫一些數值進去，但實際上不用存
                    dict_per_trading_data[ "trading_price" ] = 0
                    dict_per_trading_data[ "trading_count" ] = 0
                else:
                    dict_per_trading_data[ "trading_price" ] = item[ TradingData.TRADING_PRICE ]
                    dict_per_trading_data[ "trading_count" ] = item[ TradingData.TRADING_COUNT ]
                dict_per_trading_data[ "trading_fee_discount" ] = item[ TradingData.TRADING_FEE_DISCOUNT ]
                dict_per_trading_data[ "stock_dividend_per_share" ] = item[ TradingData.STOCK_DIVIDEND_PER_SHARE ]
                dict_per_trading_data[ "cash_dividend_per_share" ] = item[ TradingData.CASH_DIVIDEND_PER_SHARE ]
                dict_per_trading_data[ "extra_insurance_fee" ] = item[ TradingData.IS_REQUIRED_EXTRA_INSURANCE_FEE ]
                dict_per_trading_data[ "capital_reduction_per_share" ] = item[ TradingData.CAPITAL_REDUCTION_PER_SHARE ]


                export_data.append( dict_per_trading_data )
        dict_all_trading_data[ "account_name" ] = '華春'
        dict_all_trading_data[ "trading_data" ] = export_data
        list_all_company_trading_data.append( dict_all_trading_data )

        with open( file_path, 'w', encoding='utf-8' ) as f:
            json.dump( list_all_company_trading_data, f, ensure_ascii=False, indent=4 )

    def load_trading_data( self, file_path, dict_trading_data ):
        if not os.path.exists( file_path ):
            return
        with open( file_path,'r', encoding='utf-8' ) as f:
            data = json.load( f )

        for item_company in data:
            if "account_name" in item_company and \
                "trading_data" in item_company:
                for item_trading_data in item_company[ "trading_data" ]:
                    if ( "stock_number" in item_trading_data and
                        "trading_date" in item_trading_data and
                        "trading_type" in item_trading_data and
                        "trading_price" in item_trading_data and
                        "trading_count" in item_trading_data and
                        "trading_fee_discount" in item_trading_data and
                        "stock_dividend_per_share" in item_trading_data and
                        "cash_dividend_per_share" in item_trading_data and
                        "extra_insurance_fee" in item_trading_data and
                        "capital_reduction_per_share" in item_trading_data ):

                        dict_per_trading_data = Utility.generate_trading_data( item_trading_data[ "stock_number" ],                 #股票代碼
                                                                               item_trading_data[ "trading_date" ],                 #交易日期
                                                                               TradingType( item_trading_data[ "trading_type" ] ),  #交易種類
                                                                               item_trading_data[ "trading_price" ],                #交易價格
                                                                               item_trading_data[ "trading_count" ],                #交易股數
                                                                               item_trading_data[ "trading_fee_discount" ],         #手續費折扣
                                                                               item_trading_data[ "stock_dividend_per_share" ],     #每股股票股利
                                                                               item_trading_data[ "cash_dividend_per_share" ],      #每股現金股利
                                                                               item_trading_data[ "extra_insurance_fee" ],          #是否須扣除補充保費
                                                                               item_trading_data[ "capital_reduction_per_share" ] ) #每股減資金額             

                        if item_trading_data[ "stock_number" ] not in dict_trading_data:
                            dict_trading_data[ item_trading_data[ "stock_number" ] ] = [ dict_per_trading_data ]
                        else:
                            dict_trading_data[ item_trading_data[ "stock_number" ] ].append( dict_per_trading_data )

    def initialize( self ):
        self.load_UI_state()
        self.load_trading_data( g_trading_data_json_file_path, self.dict_all_stock_trading_data )
        self.process_all_trading_data()
        self.refresh_stock_list_table()

    def refresh_stock_list_table( self ):
        with QSignalBlocker( self.ui.qtStockListTableView.horizontalHeader() ):
            self.stock_list_model.clear()
            self.stock_list_model.setHorizontalHeaderLabels( g_list_stock_list_table_horizontal_header )

            list_vertical_labels = []
            for index_row,( key_stock_number, value ) in enumerate( self.dict_all_stock_trading_data.items() ):
                list_stock_name_and_type = self.dict_all_company_number_to_name_and_type[ key_stock_number ]
                str_stock_name = list_stock_name_and_type[ 0 ]
                list_vertical_labels.append( f"{key_stock_number} {str_stock_name}" )

                dict_trading_data = value[ len( value ) - 1 ] #取最後一筆交易資料，因為最後一筆交易資料的庫存等內容才是所有累計的結果
                n_accumulated_cost = dict_trading_data[ TradingData.ACCUMULATED_COST ]
                n_accumulated_inventory = dict_trading_data[ TradingData.ACCUMULATED_INVENTORY ]
                f_average_cost = round( dict_trading_data[ TradingData.AVERAGE_COST ], 3 )

                if key_stock_number in self.dict_all_company_number_to_price_info:
                    try:
                        f_stock_price = float( self.dict_all_company_number_to_price_info[ key_stock_number ] )
                        str_stock_price = format( f_stock_price, "," )
                        n_net_value = int( n_accumulated_inventory * f_stock_price )
                        str_net_value = format( n_net_value, "," )
                        n_profit = n_net_value - n_accumulated_cost
                        str_profit = format( n_profit, "," )
                        if n_profit > 0:
                            str_color = QBrush( '#FF0000' )
                        elif n_profit < 0:
                            str_color = QBrush( '#00AA00' )
                        else:
                            str_color = QBrush( '#FFFFFF' )
                    except ValueError:
                        str_stock_price = "N/A"
                        str_net_value = "N/A"
                        str_profit = "N/A"
                        str_color = QBrush( '#FFFFFF' )
                else:
                    str_stock_price = "N/A"
                    str_net_value = "N/A"
                    str_profit = "N/A"
                    str_color = QBrush( '#FFFFFF' )
                

                list_data = [ format( n_accumulated_cost, "," ),      #總成本
                            format( n_accumulated_inventory, "," ), #庫存股數
                            format( f_average_cost, "," ),          #平均成本
                            str_stock_price,                        #當前股價
                            str_net_value,                          #淨值
                            str_profit  ]                           #損益
                                
                for column, data in enumerate( list_data ):
                    standard_item = QStandardItem( data )
                    standard_item.setTextAlignment( Qt.AlignHCenter | Qt.AlignVCenter )
                    standard_item.setFlags( standard_item.flags() & ~Qt.ItemIsEditable )
                    if column == len( list_data ) - 1:
                        standard_item.setForeground( QBrush( str_color ) )
                    self.stock_list_model.setItem( index_row, column, standard_item ) 

                    
                delete_icon_item = QStandardItem("")
                delete_icon_item.setIcon( delete_icon )
                delete_icon_item.setFlags( delete_icon_item.flags() & ~Qt.ItemIsEditable )
                export_icon_item = QStandardItem("")
                export_icon_item.setIcon( export_icon )
                export_icon_item.setFlags( export_icon_item.flags() & ~Qt.ItemIsEditable )

                self.stock_list_model.setItem( index_row, len( g_list_stock_list_table_horizontal_header ) - 1, delete_icon_item )
                self.stock_list_model.setItem( index_row, len( g_list_stock_list_table_horizontal_header ) - 2, export_icon_item )

            for column in range( len( g_list_stock_list_table_horizontal_header ) ):
                if column < len( self.list_stock_list_column_width ):
                    self.ui.qtStockListTableView.setColumnWidth( column, self.list_stock_list_column_width[ column ] )
                else:
                    self.ui.qtStockListTableView.setColumnWidth( column, 100 )
                    self.list_stock_list_column_width.append( 100 )

            self.stock_list_model.setVerticalHeaderLabels( list_vertical_labels )

    def get_trading_data_header( self ):
        if self.ui.qtShow1StockRadioButton.isChecked():
            return ['年度', '日期', '交易種類', '交易價格', '交易股數', '交易金額', '手續費', '交易稅', '補充保費', '單筆總成本', '全部股票股利 /\n每股股票股利', '全部現金股利 /\n每股現金股利',
                    '累計總成本', '庫存股數', '平均成本', '編輯', '刪除' ]
        else:
            return ['年度', '日期', '交易種類', '交易價格', '交易張數', '交易金額', '手續費', '交易稅', '補充保費', '單筆總成本', '全部股票股利 /\n每股股票股利', '全部現金股利 /\n每股現金股利',
                    '累計總成本', '庫存張數', '平均成本', '編輯', '刪除' ]

    def get_per_trading_data_text_list( self, dict_per_trading_data ):
        e_trading_type = dict_per_trading_data[ TradingData.TRADING_TYPE ]
        if e_trading_type == TradingType.TEMPLATE:
            return []
        str_date = dict_per_trading_data[ TradingData.TRADING_DATE ]
        str_year = str_date.split( '-' )[ 0 ]
        if self.ui.qtROCYearRadioButton.isChecked():
            str_year = str( int( str_year ) - 1911 )

        str_month_date = str_date[ 5: ].replace( '-', '/' )
        obj_date = datetime.datetime.strptime( str_date, "%Y-%m-%d" )
        n_weekday = obj_date.weekday()
        if n_weekday == 0:
            str_weekday = "(一)"
        elif n_weekday == 1:
            str_weekday = "(二)"
        elif n_weekday == 2:
            str_weekday = "(三)"
        elif n_weekday == 3:
            str_weekday = "(四)"
        elif n_weekday == 4:
            str_weekday = "(五)"
        elif n_weekday == 5:
            str_weekday = "(六)"
        elif n_weekday == 6:
            str_weekday = "(日)"
        f_trading_price = dict_per_trading_data[ TradingData.TRADING_PRICE ]
        n_trading_count = dict_per_trading_data[ TradingData.TRADING_COUNT ]
        n_trading_value = dict_per_trading_data[ TradingData.TRADING_VALUE ]
        n_trading_fee = dict_per_trading_data[ TradingData.TRADING_FEE ]
        n_trading_tax = dict_per_trading_data[ TradingData.TRADING_TAX ]
        n_extra_insurance_fee = dict_per_trading_data[ TradingData.EXTRA_INSURANCE_FEE ]
        n_per_trading_total_cost = dict_per_trading_data[ TradingData.TRADING_COST ]
        f_stock_dividend_per_share = dict_per_trading_data[ TradingData.STOCK_DIVIDEND_PER_SHARE ]
        f_cash_dividend_per_share = dict_per_trading_data[ TradingData.CASH_DIVIDEND_PER_SHARE ]
        n_stock_dividend_gain = dict_per_trading_data[ TradingData.STOCK_DIVIDEND_GAIN ]
        n_cash_dividend_gain = dict_per_trading_data[ TradingData.CASH_DIVIDEND_GAIN ]
        n_accumulated_cost = dict_per_trading_data[ TradingData.ACCUMULATED_COST ]
        n_accumulated_inventory = dict_per_trading_data[ TradingData.ACCUMULATED_INVENTORY ]
        f_average_cost = round( dict_per_trading_data[ TradingData.AVERAGE_COST ], 3 )
        if self.ui.qtShow1StockRadioButton.isChecked():
            str_trading_count = format( n_trading_count, "," )
            str_stock_dividend_gain = format( n_stock_dividend_gain, "," )
            str_accumulated_inventory = format( n_accumulated_inventory, "," )
        else:
            f_trading_count = n_trading_count / 1000
            f_stock_dividend_gain = n_stock_dividend_gain / 1000
            f_accumulated_inventory = n_accumulated_inventory / 1000
            if f_trading_count.is_integer():
                str_trading_count = format( int( f_trading_count ), "," )
            else:
                str_trading_count = format( f_trading_count, "," )
            if f_stock_dividend_gain.is_integer():
                str_stock_dividend_gain = format( int( f_stock_dividend_gain ), "," )
            else:
                str_stock_dividend_gain = format( f_stock_dividend_gain, "," )
            if f_accumulated_inventory.is_integer():
                str_accumulated_inventory = format( int( f_accumulated_inventory ), "," )
            else:
                str_accumulated_inventory = format( f_accumulated_inventory, "," )


        if e_trading_type == TradingType.BUY:
            str_trading_type = "買進"
        elif e_trading_type == TradingType.SELL:
            n_trading_count = -n_trading_count
            n_trading_value = -n_trading_value
            n_per_trading_total_cost = -n_per_trading_total_cost
            str_trading_type = "賣出"
        elif e_trading_type == TradingType.DIVIDEND:
            str_trading_type = "股利分配"
        elif e_trading_type == TradingType.CAPITAL_REDUCTION:
            str_trading_type = "減資"

        list_data = [ str_year,                                          #交易年度
                      str_month_date + str_weekday,                      #交易日期
                      str_trading_type,                                  #交易種類
                      format( f_trading_price, "," ),                    #交易價格
                      str_trading_count,                                 #交易股數
                      format( n_trading_value, "," ),                    #交易金額
                      format( n_trading_fee, "," ),                      #手續費
                      format( n_trading_tax, "," ),                      #交易稅
                      format( n_extra_insurance_fee, "," ),                #補充保費
                      format( n_per_trading_total_cost, "," ),           #單筆總成本
                      str_stock_dividend_gain + ' / ' + str( f_stock_dividend_per_share ), #總獲得股數 / 每股股票股利
                      format( n_cash_dividend_gain, "," ) + ' / ' + str( f_cash_dividend_per_share ), #總獲得現金 / 每股現金股利
                      format( n_accumulated_cost, "," ),                 #累計總成本
                      str_accumulated_inventory,                         #庫存股數
                      format( f_average_cost, "," ) ]                    #均價
        return list_data

    def refresh_trading_data_table( self, sorted_list ):
        self.per_stock_trading_data_model.clear()
        self.per_stock_trading_data_model.setVerticalHeaderLabels( self.get_trading_data_header() )
        self.ui.qtTradingDataTableView.horizontalHeader().hide()

        if self.ui.qtFromNewToOldRadioButton.isChecked():
            loop_list = sorted_list[::-1]
            if self.ui.qtShow10RadioButton.isChecked():
                loop_list = loop_list[:10]
        else:
            loop_list = sorted_list
            if self.ui.qtShow10RadioButton.isChecked():
                loop_list = loop_list[:11]

        column = 0
        for dict_per_trading_data in loop_list:
            e_trading_type = dict_per_trading_data[ TradingData.TRADING_TYPE ]
            if e_trading_type == TradingType.TEMPLATE:
                continue

            list_data = self.get_per_trading_data_text_list( dict_per_trading_data )

            for row, data in enumerate( list_data ):
                standard_item = QStandardItem( data )
                if data == "買進":
                    standard_item.setBackground( QBrush( '#550000' ) )
                elif data == "賣出":
                    standard_item.setBackground( QBrush( '#005555' ) )
                elif data == "股利分配":
                    standard_item.setBackground( QBrush( '#555500' ) )
                elif data == "減資":
                    standard_item.setBackground( QBrush( '#6EBD61' ) )
                standard_item.setTextAlignment( Qt.AlignHCenter | Qt.AlignVCenter )
                standard_item.setFlags( standard_item.flags() & ~Qt.ItemIsEditable )
                self.per_stock_trading_data_model.setItem( row, column, standard_item ) 

            edit_icon_item = QStandardItem("")
            edit_icon_item.setIcon( edit_icon )
            edit_icon_item.setFlags( edit_icon_item.flags() & ~Qt.ItemIsEditable )
            edit_icon_item.setData( dict_per_trading_data[ TradingData.SORTED_INDEX ], Qt.UserRole )
            delete_icon_item = QStandardItem("")
            delete_icon_item.setIcon( delete_icon )
            delete_icon_item.setFlags( delete_icon_item.flags() & ~Qt.ItemIsEditable )
            delete_icon_item.setData( dict_per_trading_data[ TradingData.SORTED_INDEX ], Qt.UserRole )

            self.per_stock_trading_data_model.setItem( len( list_data ), column, edit_icon_item )
            self.per_stock_trading_data_model.setItem( len( list_data ) + 1, column, delete_icon_item )
            column += 1

    def check_internet_via_http( self, url="https://www.google.com", timeout=3):
        """
        檢測是否有網路連線（透過 HTTP 請求）
        :param url: 用於測試的 URL
        :param timeout: 超時時間（秒）
        :return: True（有網路連線）或 False（無網路連線）
        """
        try:
            response = requests.get(url, timeout=timeout)
            return response.status_code == 200
        except requests.RequestException:
            return False

    def send_get_request( self, url ):
        retries = 0
        while retries < 3:
            try:
                headers = { 'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36' }
                res = requests.get( url )
                if res.status_code == 200:
                    return res
                else:
                    print("\033[1;31mRequest failed\033[0m")
                    print(f"Status code {res.status_code}. Retrying...")
            except requests.exceptions.Timeout:
                print("\033[1;31mTimeout\033[0m")
            except requests.exceptions.TooManyRedirects:
                print("\033[1;31mTooManyRedirects\033[0m")
            
            retries += 1
            time.sleep(2)  # 等待2秒後重試
    
        raise Exception("Max retries exceeded. Failed to get a successful response.")

    def send_post_request( self, url, payload, max_retries = 3, timeout = 10 ):
        retries = 0
        while retries < max_retries:

            try:
                # 發送 POST 請求
                res = requests.post( url, data = payload, timeout=timeout )
                # 檢查回應的狀態碼，確保是成功的 2xx 系列
                if res.status_code == 200:
                    return res
                else:
                    print("\033[1;31mRequest failed\033[0m")
                    print(f"Status code {res.status_code}. Retrying...")
            except requests.exceptions.Timeout:
                print("\033[1;31mTimeout\033[0m")
            except requests.exceptions.TooManyRedirects:
                print("\033[1;31mTooManyRedirects\033[0m")

            retries += 1
            time.sleep(2)  # 等待2秒後重試
        
        raise Exception("Max retries exceeded. Failed to get a successful response.")

    def download_all_company_stock_number( self, str_date ): 
        dict_company_number_to_name = {}

        b_need_to_download = False
        if os.path.exists( g_stock_number_file_path ):
            with open( g_stock_number_file_path, 'r', encoding='utf-8' ) as f:
                data = f.readlines()
                for i, row in enumerate( data ):
                    if i == 0:
                        if row.strip() != str_date:
                            if self.check_internet_via_http(): #日期不一樣，且又有網路時才重新下載，不然就用舊的
                                b_need_to_download = True
                                break
                    else:
                        ele = row.strip().split( ',' )
                        if len( ele ) == 2:
                            b_need_to_download = True
                            break
                        else:
                            dict_company_number_to_name[ ele[ 0 ] ] = [ ele[ 1 ], ele[ 2 ] ]
        else:
            b_need_to_download = True

        if b_need_to_download:
            tds = []
            # 上市公司股票代碼
            companyNymUrl = "https://isin.twse.com.tw/isin/C_public.jsp?strMode=2"
            try:
                res = self.send_get_request( companyNymUrl )
                soup = BeautifulSoup( res.text, "lxml" )
                tr = soup.findAll( 'tr' )

                total_company_count = 0
                for raw in tr:
                    data = [ td.get_text() for td in raw.findAll("td" )]
                    if len( data ) == 7 and ( data[ 5 ] == 'ESVUFR' or 
                                              data[ 5 ] == 'CEOGBU' or
                                              data[ 5 ] == 'CEOGCU' or
                                              data[ 5 ] == 'CEOGDU' or 
                                              data[ 5 ] == 'CEOGEU' or 
                                              data[ 5 ] == 'CEOGMU' or
                                              data[ 5 ] == 'CEOJBU' or 
                                              data[ 5 ] == 'CEOJEU' or
                                              data[ 5 ] == 'CEOIBU' or
                                              data[ 5 ] == 'CEOIEU' or
                                              data[ 5 ] == 'CEOIRU' ): 
                        b_ETF = False if data[ 5 ] == 'ESVUFR' else True
                        total_company_count += 1
                        if '\u3000' in data[ 0 ]:
                            modified_data = data[ 0 ].split("\u3000")
                            if '-創' in modified_data[ 1 ]:
                                continue
                            modified_data_after_strip = [ modified_data[ 0 ].strip(), modified_data[ 1 ].strip(), b_ETF ]
                            tds.append( modified_data_after_strip )
            except Exception as e:
                pass                

            # 上櫃公司股票代碼
            companyNymUrl = "https://isin.twse.com.tw/isin/C_public.jsp?strMode=4"
            try:
                res = self.send_get_request( companyNymUrl )
                soup = BeautifulSoup( res.text, "lxml" )
                tr = soup.findAll( 'tr' )
                for raw in tr:
                    data = [ td.get_text() for td in raw.findAll("td") ]
                    if len( data ) == 7 and ( data[ 5 ] == 'ESVUFR' or 
                                              data[ 5 ] == 'CEOGBU' or
                                              data[ 5 ] == 'CEOGEU' or 
                                              data[ 5 ] == 'CEOJBU' or 
                                              data[ 5 ] == 'CEOIBU' ): 
                        b_ETF = False if data[ 5 ] == 'ESVUFR' else True
                        total_company_count += 1
                        if '\u3000' in data[ 0 ]:
                            modified_data = data[ 0 ].split("\u3000")
                            if '-創' in modified_data[ 1 ]:
                                continue
                            modified_data_after_strip = [ modified_data[ 0 ].strip(), modified_data[ 1 ].strip(), b_ETF ]
                            tds.append( modified_data_after_strip )
            except Exception as e:
                pass

            # 興櫃公司股票代碼
            companyNymUrl = "https://isin.twse.com.tw/isin/C_public.jsp?strMode=5"
            try:
                res = self.send_get_request( companyNymUrl )
                soup = BeautifulSoup( res.text, "lxml" )
                tr = soup.findAll( 'tr' )
                for raw in tr:
                    data = [ td.get_text() for td in raw.findAll("td") ]
                    if len( data ) == 7 and data[ 5 ] == 'ESVUFR': 
                        total_company_count += 1
                        if '\u3000' in data[ 0 ]:
                            modified_data = data[ 0 ].split("\u3000")
                            if '-創' in modified_data[ 1 ]:
                                continue
                            modified_data_after_strip = [ modified_data[ 0 ].strip(), modified_data[ 1 ].strip(), False ]
                            tds.append( modified_data_after_strip )
            except Exception as e:
                pass

            if len( tds ) == 0:
                return
            
            # 確保目錄存在，若不存在則遞歸創建
            os.makedirs( os.path.dirname( g_stock_number_file_path ), exist_ok = True )
            with open( g_stock_number_file_path, 'w', encoding='utf-8' ) as f:
                f.write( str_date + '\n' )
                for row in tds:
                    f.write( str( row[ 0 ] ) + ',' + str( row[ 1 ] ) + ',' + str( row[ 2 ] ) + '\n' )
                    dict_company_number_to_name[ row[ 0 ] ] = [ row[ 1 ], row[ 2 ] ]

        return dict_company_number_to_name
    
    def download_day_stock_price( self, str_date ):
        dict_company_number_to_price_info = {}
        b_need_to_download = False
        if os.path.exists( g_stock_price_file_path ):
            with open( g_stock_price_file_path, 'r', encoding='utf-8' ) as f:
                data = f.readlines()
                for i, row in enumerate( data ):
                    if i == 0:
                        if row.strip() != str_date:
                            if self.check_internet_via_http(): #日期不一樣，且又有網路時才重新下載，不然就用舊的
                                b_need_to_download = True
                    else:
                        ele = row.strip().split( ',' )
                        dict_company_number_to_price_info[ ele[ 0 ] ] = ele[ 2 ]
        else:
            b_need_to_download = True

        if b_need_to_download:
            # 上市公司股價從證交所取得
            # https://www.twse.com.tw/rwd/zh/afterTrading/MI_INDEX?date=20240912&type=ALLBUT0999&response=json&_=1726121461234
            url = 'https://www.twse.com.tw/rwd/zh/afterTrading/MI_INDEX?date=' + str_date + '&type=ALLBUT0999&response=json&_=1726121461234'
            try:
                res = self.send_get_request( url )
                soup = BeautifulSoup( res.content, 'html.parser' )

                all_stock_price = []
                json_str = soup.get_text()
                json_data = json.loads(json_str)
                if 'tables' in json_data:
                    for item in json_data['tables']:
                        if 'title' in item:
                            if '每日收盤行情' in item['title']:
                                for data in item['data']:
                                    #index 0 證券代號    "0050",
                                    #index 1 證券名稱    "元大台灣50",
                                    #index 2 成交股數    "16,337,565",
                                    #index 3 成交筆數    "15,442",
                                    #index 4 成交金額    "2,900,529,886",
                                    #index 5 開盤價      "176.10",
                                    #index 6 最高價      "178.65",
                                    #index 7 最低價      "176.10",
                                    #index 8 收盤價      "178.30",
                                    #index 9 漲跌(+/-)   "<p style= color:red>+<\u002fp>",
                                    #index 10 漲跌價差    "6.45",
                                    #index 11 最後揭示買價 "178.20",
                                    #index 12 最後揭示買量 "5",
                                    #index 13 最後揭示賣價 "178.30",
                                    #index 14 最後揭示賣量 "103",
                                    #index 15 本益比 

                                    list_stock_price = [ data[ 0 ], data[ 1 ], data[ 8 ].replace( ',', '' ) ] 
                                    all_stock_price.append( list_stock_price )
            except Exception as e:
                pass

            # 上櫃公司股價從櫃買中心取得
            # https://www.tpex.org.tw/www/zh-tw/afterTrading/dailyQuotes?date=2024%2F12%2F09&id=&response=html
            formatted_date = f"{str_date[:4]}%2F{str_date[4:6]}%2F{str_date[6:]}"
            url = 'https://www.tpex.org.tw/www/zh-tw/afterTrading/dailyQuotes?date=' + formatted_date + '&id=&response=html'
            try:
                res = self.send_get_request( url )
                
                soup = BeautifulSoup( res.text, "lxml" )
                tr = soup.findAll( 'tr' )
                for raw in tr:
                    if not raw.find( 'th' ):
                        td_elements = raw.findAll( "td" )
                        if len( td_elements ) == 19:
                            # index 0 證券代號	
                            # index 1 證券名稱	
                            # index 2 收盤	
                            # index 3 漲跌	
                            # index 4 開盤
                            # index 5 最高	
                            # index 6 最低
                            # index 7 均價	
                            # index 8 成交股數
                            # index 9 成交金額(元)
                            # index 10 成交筆數	
                            # index 11 最後買價	
                            # index 12 最後買量(千股)	2020/4/29 開始才有這筆資訊
                            # index 13 最後賣價	
                            # index 14 最後賣量(千股)   2020/4/29 開始才有這筆資訊
                            # index 15 發行股數	次日
                            # index 16 參考價	次日
                            # index 17 漲停價	次日
                            # index 18 跌停價
                            str_stock_number = td_elements[ 0 ].get_text().strip()
                            str_stock_name = td_elements[ 1 ].get_text().strip()
                            str_stock_price = td_elements[ 2 ].get_text().strip()
                            list_stock_price = [ str_stock_number, str_stock_name, str_stock_price.replace( ',', '' ) ] 
                            all_stock_price.append( list_stock_price )
            except Exception as e:
                pass    

            if len( all_stock_price ) == 0:
                print( "no data" )
                return dict_company_number_to_price_info
            
            # 確保目錄存在，若不存在則遞歸創建
            os.makedirs( os.path.dirname( g_stock_price_file_path ), exist_ok = True )
            with open( g_stock_price_file_path, 'w', encoding='utf-8' ) as f:
                f.write( str_date + '\n' )
                for row in all_stock_price:
                    f.write( str( row[ 0 ] ) + ',' + str( row[ 1 ] ) + ',' + str( row[ 2 ] ) + '\n' )
                    dict_company_number_to_price_info[ row[ 0 ] ] = str( row[ 2 ] )

        return dict_company_number_to_price_info
    
    def process_output_file_path( self, str_output_path, list_file_exist, str_folder_name, str_file_name, n_year, n_season, b_overwrite ):
        str_season = ''
        if n_season == 1 or n_season == 2 or n_season == 3 or n_season == 4:
            str_season = '_Q' + str( n_season )
        if str_output_path == None:
            str_output_path = os.path.join( g_data_dir, 'StockInventory', str_folder_name, str_file_name + str( n_year ) + str_season + '.txt' )
        # 確保目錄存在，若不存在則遞歸創建
        os.makedirs( os.path.dirname( str_output_path ), exist_ok = True )
        if b_overwrite or not os.path.exists( str_output_path ):
            list_file_exist[0] = False
        
        return str_output_path

    def download_yearly_dividend_data( self, n_year, str_date, str_output_path, b_overwrite ):
        # 假如是西元，轉成民國
        if n_year > 1990:
            n_year -= 1911

        file_exist = [ True ]
        str_output_path = self.process_output_file_path( str_output_path, file_exist, 'Dividend', 'Dividend_', n_year, 0, b_overwrite )
        if file_exist[0]:
            print("dividend file exists")
            return

        b_need_to_download = False
        if os.path.exists( str_output_path ):
            with open( str_output_path, 'r', encoding='utf-8' ) as f:
                data = f.readlines()
                for i, row in enumerate( data ):
                    if i == 0:
                        if row.strip() != str_date:
                            if self.check_internet_via_http(): #日期不一樣，且又有網路時才重新下載，不然就用舊的
                                b_need_to_download = True
        else:
            b_need_to_download = True

        if b_need_to_download:
            # 請求的 URL
            url = 'https://mops.twse.com.tw/mops/web/ajax_t108sb27'

            # POST 請求的數據
            payload = {
                # 'TYPEK': 'sii' if e_company_type2 == CompanyType2.LISTED else 'otc',
                'encodeURIComponent': '1',
                'firstin': '1',
                'off': '1',
                'step': '1',
                'year': str(n_year)
            }

            all_company_dividend = []
            try:
                for n_type in range( 2 ):
                    if n_type == 0:
                        payload[ 'TYPEK' ] = 'sii'
                    else:
                        payload[ 'TYPEK' ] = 'otc'

                    res = self.send_post_request( url, payload )

                    soup = BeautifulSoup( res.text, "lxml" )
                    tr = soup.findAll( 'tr' )
                    for raw in tr:
                        if not raw.find( 'th' ):
                            data = []
                            td_elements = raw.findAll( "td" )
                            if len( td_elements ) == 19:
                                for index, td in enumerate( td_elements ):
                                    text = td.get_text().strip()
                                    if index == 4 or index == 5 or index == 7 or index == 8 or index == 9 or index == 13 or index == 14:
                                        if text == '\xa0' or text == ''  or text == '-' or text == '--':
                                            data.append( 0 )
                                        else:
                                            number = float( text.replace( ',', '' ) ) 
                                            data.append( number )
                                    elif index == 12 or index == 15:
                                        if text == '\xa0' or text == ''  or text == '-' or text == '--':
                                            data.append( 0 )
                                        else:
                                            number = int( text.replace( ',', '' ) ) 
                                            data.append( number )
                                    else:
                                        if text == '\xa0' or text == ''  or text == '-' or text == '--':
                                            data.append( '--' )
                                        else:
                                            data.append( text )

                                all_company_dividend.append( data )
            except Exception as e:
                print(f"Final error: {e}")


            if len( all_company_dividend ) == 0:
                print( "no data" )
                return
            with open( str_output_path, 'w', encoding = 'utf-8' ) as f:
                f.write( str_date + '\n' )
                f.write( str( n_year ) + '\n' )
                f.write( '[0]公司代號,[1]公司名稱,[2]股利所屬期間,[3]權利分派基準日,[4]股票股利_盈餘轉增資配股(元/股),[5]股票股利_法定盈餘公積、資本公積轉增資配股(元/股),\
                        [6]股票股利_除權交易日,[7]現金股利_盈餘分配之股東現金股利(元/股),[8]現金股利_法定盈餘公積、資本公積發放之現金(元/股),[9]現金股利_特別股配發現金股利(元/股),\
                        [10]現金股利_除息交易日,[11]現金股利_現金股利發放日,[12]現金增資總股數(股),[13]現金增資認股比率(%),[14]現金增資認購價(元/股),\
                        [15]參加分派總股數,[16]公告日期,[17]公告時間,[18]普通股每股面額\n' )
                for row in all_company_dividend:
                    b_first = True
                    for ele in row:
                        if b_first:
                            f.write( str( ele ) )
                            b_first = False
                        else:
                            f.write( ',' + str( ele ) )
                    f.write( '\n')  # 用逗號分隔每個元素，並換行

    def download_all_yearly_dividend_data( self, n_dividend_data_start_year, str_date ):
        print( "\033[32m>>>>>>>>>>>>>>> Start to download all yearly dividend data.\033[0m" )
        current_date = datetime.datetime.today()
        n_current_year = current_date.year

        for n_year in range( n_dividend_data_start_year, n_current_year + 1 ):
            # 假如是西元，轉成民國
            if n_current_year > 1990:
                n_current_year -= 1911
            if n_year > 1990:
                n_year -= 1911

            b_overwrite = False
            file_exist = [ True ]
            str_output_path = self.process_output_file_path( None, file_exist, 'Dividend', 'Dividend_', n_year, 0, b_overwrite )
            if not file_exist[0] or n_year == n_current_year:
                self.download_yearly_dividend_data( n_year, str_date, str_output_path, True )
                print(f"Finish {n_year} yearly dividend " )

        print( "\033[32m<<<<<<<<<<<<<<< Finish downloading all yearly dividend data.\033[0m" )

    def read_yearly_dividend_raw_data( self, n_year ):
        if n_year > 1990:
            n_year -= 1911
        file_exist = [ True ]
        file_path = self.process_output_file_path( None, file_exist, 'Dividend', 'Dividend_', n_year, 0, False )
        if not os.path.exists( file_path ):
            return None
        
        list_all_company_dividend = []
        with open( file_path, 'r', encoding = 'utf-8' ) as f:
            data = f.readlines()
            for i, row in enumerate( data ):
                if i == 0 or i == 1 or i == 2:#i=0 檔案下載日期, i=1 資料年度, i=2 欄位名稱
                    continue
                else:
                    row = row.strip().split( ',' )
                    list_all_company_dividend.append( row )
            print( "Read " + 'StockDividend_Y' + str( n_year ) )

        return list_all_company_dividend

    def get_value_from_string( self, str_value ):
        if str_value == '' or str_value == '--':
            return 0
        return Decimal( str_value )

    def load_all_yearly_dividend_data( self, n_dividend_data_start_year ):
        current_date = datetime.datetime.today()
        n_current_year = current_date.year
        dict_stock_yearly_dividned = {}
        for n_year in range( n_dividend_data_start_year, n_current_year + 1 ):
            # 假如是西元，轉成民國
            if n_current_year > 1990:
                n_current_year -= 1911
            if n_year > 1990:
                n_year -= 1911
            list_yearly_dividend = self.read_yearly_dividend_raw_data( n_year )
            if list_yearly_dividend != None:
                for index, item in enumerate( list_yearly_dividend ):
                    f_stock_dividend_per_share = self.get_value_from_string( item[4] ) + self.get_value_from_string( item[5] )
                    str_stock_dividend_date = item[6]
                    f_cash_dividend_per_share = self.get_value_from_string( item[7] ) + self.get_value_from_string( item[8] )
                    str_cash_dividend_date = item[10]
                    str_cash_dividend_distribute_date = item[11]
                    str_year_month_date = ''

                    if ( f_stock_dividend_per_share != 0 and str_stock_dividend_date != '' ) and \
                       ( f_cash_dividend_per_share != 0 and str_cash_dividend_date != '' ):
                        #同時有現金股利和股票股利
                        
                        if str_stock_dividend_date == str_cash_dividend_date:
                            list_year_month_date = str_stock_dividend_date.split( '/' )
                            str_year = str( int( list_year_month_date[0] ) + 1911 )
                            str_year_month_date = str_year + '-' + list_year_month_date[1] + '-' + list_year_month_date[2]
                        else:
                            #股票股利和現金股利日期不同，理論上不應該出現
                            print("ERRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRR")
                            print("ERRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRR")
                            print("ERRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRR")
                            print("ERRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRR")
                            print("ERRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRR")
                            pass
                    elif ( f_stock_dividend_per_share != 0  and str_stock_dividend_date != '' ):
                        #只有股票股利
                        list_year_month_date = str_stock_dividend_date.split( '/' )
                        str_year = str( int( list_year_month_date[0] ) + 1911 )
                        str_year_month_date = str_year + '-' + list_year_month_date[1] + '-' + list_year_month_date[2]
                        pass
                    elif ( f_cash_dividend_per_share != 0 and str_cash_dividend_date != '' ):
                        #只有現金股利
                        list_year_month_date = str_cash_dividend_date.split( '/' )
                        str_year = str( int( list_year_month_date[0] ) + 1911 )
                        str_year_month_date = str_year + '-' + list_year_month_date[1] + '-' + list_year_month_date[2]
                        pass
                    
                    if str_year_month_date != '':
                        dict_dividend_data = Utility.generate_trading_data( item[0],                    #股票代碼
                                                                            str_year_month_date,        #交易日期
                                                                            TradingType.DIVIDEND,       #交易種類
                                                                            0,                          #交易價格                         
                                                                            0,                          #交易股數
                                                                            1,                          #手續費折扣                                   
                                                                            f_stock_dividend_per_share, #每股股票股利
                                                                            f_cash_dividend_per_share,  #每股現金股利
                                                                            False,                      #是否需扣除補充保費
                                                                            0 )                         #每股減資金額
                        dict_dividend_data[ TradingData.AUTO_DIVIDEND ] = True

                    
                    if item[0] in dict_stock_yearly_dividned:

                        dict_stock_yearly_dividned[ item[0] ].append( dict_dividend_data )
                    else:
                        dict_stock_yearly_dividned[ item[0] ] = [ dict_dividend_data ]

        return dict_stock_yearly_dividned


if __name__ == "__main__":
    app = QApplication(sys.argv)  # 創建應用程式
    app.setStyle('Fusion')
    window = MainWindow()  # 創建主窗口
    window.show()  # 顯示窗口
    sys.exit(app.exec())  # 進入事件循環